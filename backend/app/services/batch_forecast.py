"""
Batch forecast service.

Reads (Part Number, Tier 1) pairs from an uploaded Excel, runs the
forecast engine for each unique combination, and writes a result
workbook with one sheet per forecast month (12 sheets) plus a Summary.

COLUMN LAYOUT (per monthly sheet)
──────────────────────────────────
Part Number | Tier 1 | Weight (lbs) | Base Price ($) |
Quarter (Current) | Quarter (Previous) |
MC_Q ($/lb) | MC_Q-1 ($/lb) |
PPI_Q | PPI_Q-1 | PPI Factor |
CNG_Q ($/lb) | CNG_Q-1 ($/lb) |
AMS_Q ($/lb) | AMS_Q-1 ($/lb) | AMS Delta ($/lb) |
DF_c | Predicted Price ($)
"""

import io
import logging
from typing import Optional

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.models.response import ForecastResponse
from app.services.forecast_engine import ForecastEngine

logger = logging.getLogger(__name__)

# ─────────────────────────────────────────────────────────────────────────────
# STYLING
# ─────────────────────────────────────────────────────────────────────────────

_THIN        = Side(style="thin", color="BFBFBF")
_BORDER      = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
_ALT_FILL    = PatternFill("solid", fgColor="D6E4F0")
_WHITE_FILL  = PatternFill("solid", fgColor="FFFFFF")
_WARN_FILL   = PatternFill("solid", fgColor="FCE4D6")
_HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
_DATA_FONT   = Font(name="Arial", size=10)
_WARN_FONT   = Font(name="Arial", size=10, color="C00000")
_CENTER      = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT        = Alignment(horizontal="left",   vertical="center")


# ─────────────────────────────────────────────────────────────────────────────
# COLUMN DEFINITIONS
# ─────────────────────────────────────────────────────────────────────────────

def _build_columns() -> list[tuple[str, str, callable]]:
    def _ctx(mf): return mf.quarter_context
    return [
        ("Part Number",          "@",             lambda pn, t1, fr, mf: pn),
        ("Tier 1",               "@",             lambda pn, t1, fr, mf: t1),
        ("Weight\n(lbs)",        "0.00",          lambda pn, t1, fr, mf: fr.pwt_lbs),
        ("Base Price\n($)",      "$#,##0.0000",   lambda pn, t1, fr, mf: mf.base_price_used),
        ("Quarter\n(Current)",   "@",             lambda pn, t1, fr, mf: _ctx(mf).quarter_label),
        ("Quarter\n(Previous)",  "@",             lambda pn, t1, fr, mf: _ctx(mf).prev_quarter_label),
        ("MC_Q\n($/lb)",         "$0.000000",     lambda pn, t1, fr, mf: _ctx(mf).mc_q),
        ("MC_Q-1\n($/lb)",       "$0.000000",     lambda pn, t1, fr, mf: _ctx(mf).mc_q_prev),
        ("PPI_Q",                "0.000",         lambda pn, t1, fr, mf: _ctx(mf).ppi_q),
        ("PPI_Q-1",              "0.000",         lambda pn, t1, fr, mf: _ctx(mf).ppi_q_prev),
        ("PPI Factor",           "0.000000%",     lambda pn, t1, fr, mf: _ctx(mf).ppi_factor),
        ("CNG_Q\n($/lb)",        "$0.0000",       lambda pn, t1, fr, mf: _ctx(mf).cng_q),
        ("CNG_Q-1\n($/lb)",      "$0.0000",       lambda pn, t1, fr, mf: _ctx(mf).cng_q_prev),
        ("AMS_Q\n($/lb)",        "$0.000000",     lambda pn, t1, fr, mf: _ctx(mf).ams_q),
        ("AMS_Q-1\n($/lb)",      "$0.000000",     lambda pn, t1, fr, mf: _ctx(mf).ams_q_prev),
        ("AMS Delta\n($/lb)",    "$0.000000",     lambda pn, t1, fr, mf: _ctx(mf).ams_delta),
        ("DF_c",                 "0.00",          lambda pn, t1, fr, mf: mf.df_c),
        ("Predicted Price\n($)", "$#,##0.0000",   lambda pn, t1, fr, mf: mf.predicted_price),
    ]


COLUMNS = _build_columns()
N_COLS  = len(COLUMNS)


# ─────────────────────────────────────────────────────────────────────────────
# CELL HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def _write_header(ws, col_idx: int, label: str) -> None:
    cell = ws.cell(row=1, column=col_idx, value=label)
    cell.font = _HEADER_FONT; cell.fill = _HEADER_FILL
    cell.alignment = _CENTER; cell.border = _BORDER


def _write_data(ws, row: int, col: int, value, fmt: str, fill, left=False) -> None:
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = _DATA_FONT; cell.fill = fill; cell.border = _BORDER
    cell.number_format = fmt
    cell.alignment = _LEFT if left else _CENTER


def _write_error_row(ws, row: int, part_number: str, tier_1: str, error_msg: str, fill) -> None:
    for col, val in [(1, part_number), (2, tier_1)]:
        c = ws.cell(row=row, column=col, value=val)
        c.font = _WARN_FONT; c.fill = _WARN_FILL
        c.border = _BORDER; c.alignment = _LEFT

    err = ws.cell(row=row, column=3, value=f"ERROR: {error_msg}")
    err.font = _WARN_FONT; err.fill = _WARN_FILL
    err.border = _BORDER; err.alignment = _LEFT
    ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=N_COLS)


def _set_col_widths(ws) -> None:
    widths = [22, 22, 10, 14, 14, 14, 13, 13, 10, 10, 13, 13, 13, 13, 13, 13, 8, 18]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 36


# ─────────────────────────────────────────────────────────────────────────────
# INPUT READER
# ─────────────────────────────────────────────────────────────────────────────

def read_parts_from_upload(file_bytes: bytes) -> list[tuple[str, str]]:
    """
    Read (Part Number, Tier 1) pairs from the uploaded Excel file.

    Expects columns named 'Part Number' and 'Tier 1' (case-insensitive).
    Returns a deduplicated list of (part_number, tier_1) tuples,
    preserving input order.
    """
    df = pd.read_excel(
        io.BytesIO(file_bytes),
        engine="openpyxl",
        dtype=str,
        keep_default_na=False,
        na_values=[""],
    )

    cols_lower = {c.strip().lower(): c for c in df.columns}

    if "part number" not in cols_lower:
        raise ValueError(
            f"Input Excel must have a 'Part Number' column. "
            f"Found: {list(df.columns)}"
        )
    if "tier 1" not in cols_lower:
        raise ValueError(
            f"Input Excel must have a 'Tier 1' column. "
            f"Found: {list(df.columns)}"
        )

    pn_col = cols_lower["part number"]
    t1_col = cols_lower["tier 1"]

    seen: set[tuple[str, str]] = set()
    pairs: list[tuple[str, str]] = []
    for _, row in df.iterrows():
        pn = str(row[pn_col]).strip()
        t1 = str(row[t1_col]).strip()
        if pn and t1 and (pn, t1) not in seen:
            seen.add((pn, t1))
            pairs.append((pn, t1))

    logger.info("Read %d unique (part, tier_1) pairs from uploaded file", len(pairs))
    return pairs


# ─────────────────────────────────────────────────────────────────────────────
# OUTPUT BUILDER
# ─────────────────────────────────────────────────────────────────────────────

def _run_forecasts(
    part_tier_pairs: list[tuple[str, str]],
    engine: ForecastEngine,
) -> tuple[dict, list[tuple[str, str]]]:
    """
    Run the engine for every (part, tier_1) pair and return
    (results, month_labels). Failed rows are stored as the raised Exception.
    """
    results: dict[tuple[str, str], ForecastResponse | Exception] = {}

    for pn, t1 in part_tier_pairs:
        try:
            results[(pn, t1)] = engine.forecast(part_number=pn, tier_1=t1)
            logger.debug("Forecast OK: %s / %s", pn, t1)
        except Exception as exc:
            results[(pn, t1)] = exc
            logger.warning("Forecast failed for %s / %s: %s", pn, t1, exc)

    month_labels: list[tuple[str, str]] = []
    for r in results.values():
        if isinstance(r, ForecastResponse):
            month_labels = [(f.year_month, f.month_label) for f in r.forecasts]
            break

    if not month_labels:
        raise ValueError("All parts failed forecasting — cannot generate output workbook.")

    return results, month_labels


def _workbook_from_results(
    part_tier_pairs: list[tuple[str, str]],
    results: dict,
    month_labels: list[tuple[str, str]],
) -> bytes:
    """Serialise precomputed forecast results into a styled .xlsx workbook."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for year_month, month_label in month_labels:
        ws = wb.create_sheet(title=month_label[:31])
        ws.freeze_panes = "A2"

        for ci, (label, _, _) in enumerate(COLUMNS, 1):
            _write_header(ws, ci, label)

        for ri, (pn, t1) in enumerate(part_tier_pairs, 2):
            fill   = _ALT_FILL if ri % 2 == 0 else _WHITE_FILL
            result = results[(pn, t1)]

            if isinstance(result, Exception):
                _write_error_row(ws, ri, pn, t1, str(result), fill)
                continue

            mf = next((f for f in result.forecasts if f.year_month == year_month), None)
            if mf is None:
                _write_error_row(ws, ri, pn, t1, f"No forecast data for {year_month}", fill)
                continue

            for ci, (_, fmt, extractor) in enumerate(COLUMNS, 1):
                left = ci in (1, 2)
                try:
                    value = extractor(pn, t1, result, mf)
                except Exception as ex:
                    value = f"ERR: {ex}"
                _write_data(ws, ri, ci, value, fmt, fill, left=left)

        _set_col_widths(ws)

    # ── Step 4: summary sheet ─────────────────────────────────────────────
    _build_summary_sheet(wb, part_tier_pairs, results, month_labels)

    # ── serialise ─────────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def build_forecast_workbook(
    part_tier_pairs: list[tuple[str, str]],
    engine: ForecastEngine,
) -> bytes:
    """
    Run forecasts for all (part_number, tier_1) pairs and build output workbook.

    Failed rows show an ERROR message instead of stopping the batch.
    Returns raw bytes of the generated .xlsx workbook.
    """
    results, month_labels = _run_forecasts(part_tier_pairs, engine)
    return _workbook_from_results(part_tier_pairs, results, month_labels)


# ─────────────────────────────────────────────────────────────────────────────
# JSON PREVIEW (for in-UI rendering)
# ─────────────────────────────────────────────────────────────────────────────

def _fmt(value, fmt: str) -> str:
    """Render a raw value into a display string mirroring the Excel number format."""
    if value is None:
        return ""
    if not isinstance(value, (int, float)) or fmt == "@":
        return str(value)
    if fmt.endswith("%"):
        decimals = fmt[fmt.index(".") + 1: fmt.index("%")].count("0") if "." in fmt else 0
        return f"{value * 100:.{decimals}f}%"
    prefix = "$" if fmt.startswith("$") else ""
    comma = "," if "#,##0" in fmt else ""
    decimals = fmt.split(".")[1].count("0") if "." in fmt else 0
    return f"{prefix}{value:{comma}.{decimals}f}"


def build_forecast_bundle(
    part_tier_pairs: list[tuple[str, str]],
    engine: ForecastEngine,
) -> dict:
    """
    Run forecasts once and return both a JSON-friendly preview of every sheet
    and the raw workbook bytes (so the UI can render + download without
    re-processing).

    Returns: { "sheets": [ {name, columns, rows}, ... ], "workbook_bytes": bytes }
    """
    results, month_labels = _run_forecasts(part_tier_pairs, engine)

    sheets: list[dict] = []

    # ── Summary sheet ─────────────────────────────────────────────────────
    summary_cols = ["Part Number", "Tier 1", "Weight (lbs)", "Base Price ($)"] + \
                   [label for _, label in month_labels]
    summary_rows: list[list[str]] = []
    for pn, t1 in part_tier_pairs:
        result = results[(pn, t1)]
        if isinstance(result, Exception):
            row = [pn, t1, f"ERROR: {result}"] + [""] * (len(summary_cols) - 3)
            summary_rows.append(row)
            continue
        row = [
            pn, t1,
            _fmt(result.pwt_lbs, "0.00"),
            _fmt(result.base_price, "$#,##0.0000"),
        ]
        for year_month, _ in month_labels:
            mf = next((f for f in result.forecasts if f.year_month == year_month), None)
            row.append(_fmt(mf.predicted_price, "$#,##0.0000") if mf else "N/A")
        summary_rows.append(row)
    sheets.append({"name": "Summary", "columns": summary_cols, "rows": summary_rows})

    # ── Monthly sheets ────────────────────────────────────────────────────
    monthly_cols = [label.replace("\n", " ") for label, _, _ in COLUMNS]
    for year_month, month_label in month_labels:
        rows: list[list[str]] = []
        for pn, t1 in part_tier_pairs:
            result = results[(pn, t1)]
            if isinstance(result, Exception):
                rows.append([pn, t1, f"ERROR: {result}"] + [""] * (N_COLS - 3))
                continue
            mf = next((f for f in result.forecasts if f.year_month == year_month), None)
            if mf is None:
                rows.append([pn, t1, f"No data for {year_month}"] + [""] * (N_COLS - 3))
                continue
            row = []
            for _, fmt, extractor in COLUMNS:
                try:
                    row.append(_fmt(extractor(pn, t1, result, mf), fmt))
                except Exception as ex:
                    row.append(f"ERR: {ex}")
            rows.append(row)
        sheets.append({"name": month_label, "columns": monthly_cols, "rows": rows})

    workbook_bytes = _workbook_from_results(part_tier_pairs, results, month_labels)

    return {"sheets": sheets, "workbook_bytes": workbook_bytes}


def _build_summary_sheet(
    wb, part_tier_pairs, results, month_labels
) -> None:
    ws = wb.create_sheet(title="Summary", index=0)
    ws.freeze_panes = "E2"

    fixed   = ["Part Number", "Tier 1", "Weight (lbs)", "Base Price ($)"]
    monthly = [label for _, label in month_labels]

    for ci, h in enumerate(fixed + monthly, 1):
        _write_header(ws, ci, h)
    ws.row_dimensions[1].height = 36

    for ri, (pn, t1) in enumerate(part_tier_pairs, 2):
        fill   = _ALT_FILL if ri % 2 == 0 else _WHITE_FILL
        result = results[(pn, t1)]

        if isinstance(result, Exception):
            _write_error_row(ws, ri, pn, t1, str(result), fill)
            continue

        _write_data(ws, ri, 1, pn,               "@",            fill, left=True)
        _write_data(ws, ri, 2, t1,               "@",            fill, left=True)
        _write_data(ws, ri, 3, result.pwt_lbs,   "0.00",         fill)
        _write_data(ws, ri, 4, result.base_price, "$#,##0.0000", fill)

        for mi, (year_month, _) in enumerate(month_labels):
            mf  = next((f for f in result.forecasts if f.year_month == year_month), None)
            val = mf.predicted_price if mf else "N/A"
            _write_data(ws, ri, 5 + mi, val, "$#,##0.0000", fill)

    # Column widths
    for col, width in zip("ABCD", [22, 22, 12, 16]):
        ws.column_dimensions[col].width = width
    for i in range(len(month_labels)):
        ws.column_dimensions[get_column_letter(5 + i)].width = 18