"""Builds aluminium_data.xlsx (Parts, LME, Midwest, PPI, CNG sheets) from
app/data/real/*.csv so the Excel-backed forecast engine has its data file.
Run: python build_aluminium_xlsx.py"""
import csv
from pathlib import Path

import openpyxl

BASE = Path(__file__).parent
REAL = BASE / "app" / "data" / "real"
OUT = BASE / "aluminium_data.xlsx"

SHEET_MAP = {
    "LME": ("lme.csv", "LME Price ($/lb)"),
    "Midwest": ("midwest_premium.csv", "Midwest Premium ($/lb)"),
    "PPI": ("labour.csv", "PPI Index (dimensionless)"),
    "CNG": ("gas.csv", "CNG Cost ($/lb)"),
}

wb = openpyxl.Workbook()
wb.remove(wb.active)

ws = wb.create_sheet("Parts")
with open(REAL / "parts.csv") as f:
    for row in csv.reader(f):
        ws.append(row)
for r in ws.iter_rows(min_row=2):
    r[2].value = float(r[2].value)
    r[3].value = float(r[3].value)

for sheet, (fname, header) in SHEET_MAP.items():
    ws = wb.create_sheet(sheet)
    ws.append(["Month (YYYY-MM)", header])
    with open(REAL / fname) as f:
        for row in csv.DictReader(f):
            ws.append([row["month"], float(row["value"])])

wb.save(OUT)
print(f"wrote {OUT}")
