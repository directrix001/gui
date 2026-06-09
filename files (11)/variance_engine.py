"""
variance_engine.py  —  v4.2
────────────────────────────
stdlib only — no lxml, no third-party XML library.

TIER 1  (Windows + Excel installed)
  xlwings COM: Excel's own Worksheet.Copy() keeps all formula refs internal.
  BreakLink() removes template refs. CalculateFull() recalcs before save.

TIER 2  (fallback — no Excel / xlwings not installed)
  zip-level XML surgery:
    • Unzip both .xlsx files (they are plain zip archives).
    • Copy raw worksheet XML byte-for-byte — preserves charts/images/formatting.
    • Register new sheets in workbook.xml, workbook.xml.rels, [Content_Types].xml.
    • Strip every '[TemplateName.xlsx]' prefix from formula <f> elements.
    • Remove <externalReferences> so Excel never prompts "Update Links?".
    • Repack into a clean .xlsx.

Key fixes in v4.2 vs v4.1
  • Removed ET.register_namespace("") — registering the default namespace
    causes ElementTree to emit ns0: prefixes, corrupting workbook.xml.
  • Fixed relationship Type URI (was broken by string-slicing _NS_R).
  • workbook.xml and rels files are now written via raw string replacement
    to avoid ElementTree re-serialisation changing namespace declarations.
  • [Content_Types].xml patched the same safe way.
"""

import os
import re
import shutil
import zipfile
import tempfile
from datetime import datetime

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill


# ═══════════════════════════════════════════════════════════════
#  CONFIGURATION — set explicitly if auto-detect picks wrong sheet
# ═══════════════════════════════════════════════════════════════
MTD_SHEET_NAME: str | None = None
YTD_SHEET_NAME: str | None = None


# ── Relationship type constants ──────────────────────────────────────────────
_REL_WORKSHEET = (
    "http://schemas.openxmlformats.org/officeDocument/2006/"
    "relationships/worksheet"
)
_NS_MAIN = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
_NS_R    = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
_NS_CT   = "http://schemas.openxmlformats.org/package/2006/content-types"
_NS_RELS = "http://schemas.openxmlformats.org/package/2006/relationships"


# ═══════════════════════════════════════════════════════════════
#  MAIN ENTRY POINT
# ═══════════════════════════════════════════════════════════════

def run_variance(args: dict) -> str:
    """
    Produce one output .xlsx:
      base  = master file copied to output folder
      added = MTD + YTD tabs from template for every scenario pair
      last  = Input Log audit tab
    Returns full path of saved file.
    """
    pairs         = args["scenario_pairs"]
    quarter       = args["quarter"]
    month         = args["month"]
    master_file   = args["master_file"]
    template_file = args["template_file"]
    out_folder    = args["output_folder"]
    ts            = args["timestamp"]

    os.makedirs(out_folder, exist_ok=True)

    if not os.path.isfile(template_file):
        raise FileNotFoundError(f"Template file not found:\n{template_file}")
    if not os.path.isfile(master_file):
        raise FileNotFoundError(f"Master file not found:\n{master_file}")

    mtd_name, ytd_name = _resolve_template_sheets(template_file)

    q_label     = quarter.split()[0]
    output_path = os.path.join(
        out_folder, f"Variance__{q_label}_{month}__v{ts}.xlsx")
    shutil.copy2(master_file, output_path)

    # TIER 1 — Excel COM via xlwings (best quality, Windows only)
    ok = _try_xlwings_copy(output_path, template_file, pairs, mtd_name, ytd_name)

    if not ok:
        # TIER 2 — zip / XML surgery (cross-platform, no Excel needed)
        print("[engine] Using zip-level XML copy.")
        _zip_copy_sheets(output_path, template_file, pairs, mtd_name, ytd_name)

    # Append audit log (plain data — no formulas, safe to write with openpyxl)
    out_wb = load_workbook(output_path, data_only=False, keep_vba=False)
    if "Input Log" in out_wb.sheetnames:
        del out_wb["Input Log"]
    _write_input_log(out_wb.create_sheet("Input Log"), args, pairs)
    out_wb.save(output_path)
    out_wb.close()

    print(f"[engine] Done: {os.path.basename(output_path)}")
    return output_path


# ═══════════════════════════════════════════════════════════════
#  TIER 1 — xlwings COM  (Excel-native copy)
# ═══════════════════════════════════════════════════════════════

def _try_xlwings_copy(output_path, template_file,
                      pairs, mtd_name, ytd_name) -> bool:
    try:
        import xlwings as xw
    except ImportError:
        print("[engine] xlwings not installed — zip fallback.")
        return False

    app = None
    try:
        app = xw.App(visible=False, add_book=False)
        app.display_alerts  = False
        app.screen_updating = False

        tpl_wb = app.books.open(os.path.abspath(template_file))
        out_wb = app.books.open(os.path.abspath(output_path))

        for sc1, sc2 in pairs:
            lbl = f"{sc1} vs {sc2}"
            for src_name, new_name in [
                (mtd_name, f"MTD {lbl}"[:31]),
                (ytd_name, f"YTD {lbl}"[:31]),
            ]:
                # Remove existing tab with same name
                for sh in out_wb.sheets:
                    if sh.name == new_name:
                        sh.delete()
                        break
                # Excel's own Copy() — all formula refs stay internal
                tpl_wb.sheets[src_name].api.Copy(
                    After=out_wb.sheets[-1].api)
                out_wb.sheets[-1].name = new_name
                print(f"[engine][xlwings] '{src_name}' → '{new_name}'")

        _xlwings_break_links(out_wb, template_file)
        app.api.CalculateFull()
        out_wb.save()
        out_wb.close()
        tpl_wb.close()
        app.quit()
        return True

    except Exception as exc:
        print(f"[engine] xlwings failed ({exc}) — zip fallback.")
        try:
            if app:
                app.quit()
        except Exception:
            pass
        return False


def _xlwings_break_links(wb_xw, template_file: str):
    """Sever external links back to the template workbook."""
    try:
        links = wb_xw.api.LinkSources(1)   # 1 = xlExcelLinks
        if not links:
            return
        tpl_lower = os.path.basename(template_file).lower()
        for link in links:
            if tpl_lower in str(link).lower():
                wb_xw.api.BreakLink(Name=link, Type=1)
                print(f"[engine][xlwings] Broke link: {link}")
    except Exception as e:
        print(f"[engine][xlwings] BreakLink skipped: {e}")


# ═══════════════════════════════════════════════════════════════
#  TIER 2 — zip / XML surgery  (stdlib only)
# ═══════════════════════════════════════════════════════════════

def _zip_copy_sheets(output_path: str, template_file: str,
                     pairs: list, mtd_name: str, ytd_name: str):
    """
    Copy MTD + YTD sheets from template into output at the raw zip/XML level.
    All modifications use safe string/regex operations on the raw bytes —
    we never re-serialise XML through ElementTree (which rewrites namespace
    declarations and can corrupt the file).
    """
    tpl_base    = os.path.basename(template_file)            # e.g. "Template.xlsx"
    tpl_base_nx = os.path.splitext(tpl_base)[0]             # e.g. "Template"

    with tempfile.TemporaryDirectory() as tmpdir:
        out_dir = os.path.join(tmpdir, "out")
        tpl_dir = os.path.join(tmpdir, "tpl")

        with zipfile.ZipFile(output_path,   "r") as z: z.extractall(out_dir)
        with zipfile.ZipFile(template_file, "r") as z: z.extractall(tpl_dir)

        # ── Discover sheet name → file mapping ──────────────────────────
        tpl_sheet_map = _read_sheet_map(tpl_dir)  # {name: "worksheets/sheetN.xml"}
        out_sheet_map = _read_sheet_map(out_dir)

        next_sh  = _next_sheet_index(out_dir)
        next_rid = _next_rid_index(out_dir)

        sheets_to_copy = []
        for sc1, sc2 in pairs:
            lbl = f"{sc1} vs {sc2}"
            sheets_to_copy.append((mtd_name, f"MTD {lbl}"[:31]))
            sheets_to_copy.append((ytd_name, f"YTD {lbl}"[:31]))

        for tpl_sname, new_tab in sheets_to_copy:
            if tpl_sname not in tpl_sheet_map:
                raise ValueError(
                    f"Sheet '{tpl_sname}' not found in template.\n"
                    f"Available sheets: {list(tpl_sheet_map.keys())}")

            tpl_ws_file   = tpl_sheet_map[tpl_sname]   # "worksheets/sheet2.xml"
            new_ws_file   = f"worksheets/sheet{next_sh}.xml"
            new_ws_path   = os.path.join(out_dir, "xl", new_ws_file)
            new_rid       = f"rId{next_rid}"

            # 1. Copy worksheet XML and strip external-link prefixes
            src_ws = os.path.join(tpl_dir, "xl", tpl_ws_file)
            os.makedirs(os.path.dirname(new_ws_path), exist_ok=True)
            shutil.copy2(src_ws, new_ws_path)
            _strip_external_links(new_ws_path, tpl_base, tpl_base_nx)

            # 2. Copy worksheet .rels + all dependencies (charts/images/drawings)
            tpl_ws_rels_path = os.path.join(
                tpl_dir, "xl", "worksheets", "_rels",
                os.path.basename(tpl_ws_file) + ".rels")
            if os.path.isfile(tpl_ws_rels_path):
                out_ws_rels_dir = os.path.join(out_dir, "xl", "worksheets", "_rels")
                os.makedirs(out_ws_rels_dir, exist_ok=True)
                out_ws_rels_path = os.path.join(
                    out_ws_rels_dir, f"sheet{next_sh}.xml.rels")
                shutil.copy2(tpl_ws_rels_path, out_ws_rels_path)
                _copy_dependencies(tpl_ws_rels_path, tpl_dir, out_dir, tpl_ws_file)

            # 3. Register in workbook.xml  (raw string insertion)
            _patch_workbook_xml(out_dir, new_tab, str(next_sh), new_rid)

            # 4. Register in workbook.xml.rels  (raw string insertion)
            _patch_workbook_rels(out_dir, new_rid, new_ws_file)

            # 5. Register in [Content_Types].xml  (raw string insertion)
            _patch_content_types(out_dir, f"/xl/{new_ws_file}")

            next_sh  += 1
            next_rid += 1
            print(f"[engine][zip] '{tpl_sname}' → '{new_tab}'")

        # 6. Remove stale <externalReferences> block from workbook.xml
        _remove_external_refs(out_dir, tpl_base, tpl_base_nx)

        # 7. Repack into .xlsx
        _repack_zip(out_dir, output_path)


# ── Sheet map (parse workbook.xml + rels via regex — no ET re-serialise) ─────

def _read_sheet_map(base_dir: str) -> dict:
    """
    Returns {sheet_name: "worksheets/sheetN.xml"} by reading
    workbook.xml and workbook.xml.rels with regex (safe, no namespace issues).
    """
    wb_path   = os.path.join(base_dir, "xl", "workbook.xml")
    rels_path = os.path.join(base_dir, "xl", "_rels", "workbook.xml.rels")

    wb_text   = _read_text(wb_path)
    rels_text = _read_text(rels_path)

    # Build rId → target from rels
    rid_to_target = {}
    for m in re.finditer(
            r'<Relationship\b[^>]*\bId=["\']([^"\']+)["\'][^>]*\bTarget=["\']([^"\']+)["\']',
            rels_text):
        rid_to_target[m.group(1)] = m.group(2)
    # Also handle reversed attribute order
    for m in re.finditer(
            r'<Relationship\b[^>]*\bTarget=["\']([^"\']+)["\'][^>]*\bId=["\']([^"\']+)["\']',
            rels_text):
        rid_to_target[m.group(2)] = m.group(1)

    # Parse <sheet name="..." r:id="..."> from workbook.xml
    result = {}
    for m in re.finditer(
            r'<sheet\b[^>]*\bname=["\']([^"\']+)["\'][^>]*/?>',
            wb_text):
        tag  = m.group(0)
        name = m.group(1)
        rid_m = re.search(r'r:id=["\']([^"\']+)["\']', tag)
        if rid_m:
            rid    = rid_m.group(1)
            target = rid_to_target.get(rid, "")
            result[name] = target
    return result


# ── Index helpers ─────────────────────────────────────────────────────────────

def _next_sheet_index(out_dir: str) -> int:
    ws_dir = os.path.join(out_dir, "xl", "worksheets")
    if not os.path.isdir(ws_dir):
        return 1
    nums = []
    for f in os.listdir(ws_dir):
        m = re.search(r"^sheet(\d+)\.xml$", f)
        if m:
            nums.append(int(m.group(1)))
    return max(nums, default=0) + 1


def _next_rid_index(out_dir: str) -> int:
    rels_path = os.path.join(out_dir, "xl", "_rels", "workbook.xml.rels")
    text = _read_text(rels_path)
    nums = [int(m.group(1)) for m in re.finditer(r'\brId(\d+)\b', text)]
    return max(nums, default=0) + 1


# ── Raw-string XML patchers  (no ElementTree re-serialise) ───────────────────

def _patch_workbook_xml(out_dir: str, sheet_name: str,
                        sheet_idx: str, r_id: str):
    """
    Insert a new <sheet> element into workbook.xml.
    Finds the closing </sheets> tag and inserts before it.
    """
    path = os.path.join(out_dir, "xl", "workbook.xml")
    text = _read_text(path)

    # Determine next sheetId
    ids  = [int(m.group(1)) for m in re.finditer(r'\bsheetId=["\'](\d+)["\']', text)]
    next_id = max(ids, default=0) + 1

    # Escape XML special chars in sheet name
    safe_name = (sheet_name
                 .replace("&", "&amp;")
                 .replace('"', "&quot;")
                 .replace("<", "&lt;")
                 .replace(">", "&gt;"))

    new_el = (f'<sheet name="{safe_name}" sheetId="{next_id}" '
              f'r:id="{r_id}"/>')

    # Insert before </sheets>
    if "</sheets>" in text:
        text = text.replace("</sheets>", new_el + "</sheets>", 1)
    else:
        # Fallback: insert before </workbook>
        text = text.replace("</workbook>", f"<sheets>{new_el}</sheets></workbook>", 1)

    _write_text(path, text)


def _patch_workbook_rels(out_dir: str, r_id: str, ws_file: str):
    """
    Insert a new <Relationship> into workbook.xml.rels before </Relationships>.
    """
    path = os.path.join(out_dir, "xl", "_rels", "workbook.xml.rels")
    text = _read_text(path)

    new_rel = (f'<Relationship Id="{r_id}" '
               f'Type="{_REL_WORKSHEET}" '
               f'Target="{ws_file}"/>')

    text = text.replace("</Relationships>", new_rel + "</Relationships>", 1)
    _write_text(path, text)


def _patch_content_types(out_dir: str, part_name: str):
    """
    Add an <Override> entry to [Content_Types].xml if not already present.
    """
    path = os.path.join(out_dir, "[Content_Types].xml")
    text = _read_text(path)

    if part_name in text:
        return   # already registered

    new_ov = (f'<Override PartName="{part_name}" '
              f'ContentType="application/vnd.openxmlformats-officedocument'
              f'.spreadsheetml.worksheet+xml"/>')

    text = text.replace("</Types>", new_ov + "</Types>", 1)
    _write_text(path, text)


def _remove_external_refs(out_dir: str, tpl_base: str, tpl_base_nx: str):
    """
    Remove <externalReferences>…</externalReferences> block and any
    <definedName> entries referencing the template from workbook.xml.
    Also removes <externalLink> entries from workbook.xml.rels.
    """
    wb_path = os.path.join(out_dir, "xl", "workbook.xml")
    text    = _read_text(wb_path)

    # Remove the whole externalReferences block
    text = re.sub(
        r'<externalReferences\b[^>]*>.*?</externalReferences>',
        '', text, flags=re.DOTALL | re.IGNORECASE)
    text = re.sub(
        r'<externalReferences\b[^/]*/?>',
        '', text, flags=re.IGNORECASE)

    # Remove definedNames that reference the template file
    tpl_lo = re.escape(tpl_base.lower())
    nox_lo = re.escape(tpl_base_nx.lower())

    def _should_remove_dn(dn_match):
        val = dn_match.group(0).lower()
        return tpl_lo in val or nox_lo in val

    text = re.sub(
        r'<definedName\b[^>]*>.*?</definedName>',
        lambda m: '' if _should_remove_dn(m) else m.group(0),
        text, flags=re.DOTALL | re.IGNORECASE)

    _write_text(wb_path, text)

    # Remove externalLink relationships from workbook.xml.rels
    rels_path = os.path.join(out_dir, "xl", "_rels", "workbook.xml.rels")
    rels_text = _read_text(rels_path)
    rels_text = re.sub(
        r'<Relationship\b[^>]*externalLink[^>]*/?>',
        '', rels_text, flags=re.IGNORECASE)
    _write_text(rels_path, rels_text)


# ── Strip external-link prefixes from worksheet XML ──────────────────────────

def _strip_external_links(ws_path: str, tpl_base: str, tpl_base_nx: str):
    """
    Remove '[Template.xlsx]' prefixes from every formula <f>…</f> in the
    worksheet XML.  Works on raw bytes — no XML re-serialisation.
    """
    pats = [
        re.compile(r'\[' + re.escape(tpl_base)           + r'\]', re.IGNORECASE),
        re.compile(r'\[' + re.escape(tpl_base_nx) + r'\.xlsx\]', re.IGNORECASE),
        re.compile(r'\[' + re.escape(tpl_base_nx) + r'\.xlsm\]', re.IGNORECASE),
        re.compile(r'\[' + re.escape(tpl_base_nx)         + r'\]', re.IGNORECASE),
    ]

    def _clean(s: str) -> str:
        for p in pats:
            s = p.sub("", s)
        return s

    text = _read_text(ws_path)
    # Replace content between <f …> … </f> tags
    text = re.sub(
        r'(<f(?:\s[^>]*)?>)(.*?)(</f>)',
        lambda m: m.group(1) + _clean(m.group(2)) + m.group(3),
        text, flags=re.DOTALL
    )
    _write_text(ws_path, text)


# ── Copy drawing / chart / image dependencies ─────────────────────────────────

def _copy_dependencies(rels_path: str, tpl_dir: str,
                        out_dir: str, tpl_ws_file: str):
    """
    Read worksheet .rels and copy every referenced resource
    (drawings, charts, images, vmlDrawing) from template to output.
    """
    text      = _read_text(rels_path)
    ws_subdir = os.path.dirname(tpl_ws_file)   # "worksheets"

    for m in re.finditer(r'Target=["\']([^"\']+)["\']', text):
        target  = m.group(1)
        src_rel = os.path.normpath(
            os.path.join("xl", ws_subdir, target)
        ).replace("\\", "/")

        src_abs = os.path.join(tpl_dir, src_rel)
        dst_abs = os.path.join(out_dir, src_rel)

        if os.path.isfile(src_abs) and not os.path.isfile(dst_abs):
            os.makedirs(os.path.dirname(dst_abs), exist_ok=True)
            shutil.copy2(src_abs, dst_abs)

            # Copy the resource's own .rels if present
            dep_rels_src = os.path.join(
                os.path.dirname(src_abs), "_rels",
                os.path.basename(src_abs) + ".rels")
            if os.path.isfile(dep_rels_src):
                dep_rels_dst = os.path.join(
                    os.path.dirname(dst_abs), "_rels",
                    os.path.basename(dst_abs) + ".rels")
                os.makedirs(os.path.dirname(dep_rels_dst), exist_ok=True)
                shutil.copy2(dep_rels_src, dep_rels_dst)


# ── Repack directory → .xlsx ──────────────────────────────────────────────────

def _repack_zip(src_dir: str, out_path: str):
    """Walk src_dir and repack everything into out_path as a ZIP/XLSX."""
    tmp = out_path + ".tmp_repack"
    with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
        for root_dir, _dirs, files in os.walk(src_dir):
            # [Content_Types].xml must be the very first entry in the zip
            if root_dir == src_dir and "[Content_Types].xml" in files:
                files = list(files)
                files.remove("[Content_Types].xml")
                files.insert(0, "[Content_Types].xml")
            for fname in files:
                abs_path = os.path.join(root_dir, fname)
                arc_name = os.path.relpath(abs_path, src_dir).replace("\\", "/")
                zout.write(abs_path, arc_name)
    os.replace(tmp, out_path)


# ── Low-level text read / write ───────────────────────────────────────────────

def _read_text(path: str) -> str:
    with open(path, "rb") as f:
        raw = f.read()
    # Detect encoding from XML declaration if present
    enc = "utf-8"
    m = re.match(rb'<\?xml[^>]+encoding=["\']([^"\']+)["\']', raw[:80])
    if m:
        enc = m.group(1).decode("ascii")
    return raw.decode(enc, errors="replace")


def _write_text(path: str, text: str):
    with open(path, "wb") as f:
        f.write(text.encode("utf-8"))


# ═══════════════════════════════════════════════════════════════
#  TEMPLATE SHEET AUTO-DETECT
# ═══════════════════════════════════════════════════════════════

def _resolve_template_sheets(tpl_path: str):
    wb    = load_workbook(tpl_path, read_only=True, data_only=True)
    names = wb.sheetnames
    wb.close()

    mtd = MTD_SHEET_NAME or next(
        (n for n in names if "mtd" in n.lower()), None)
    ytd = YTD_SHEET_NAME or next(
        (n for n in names if "ytd" in n.lower()), None)

    if mtd is None:
        raise ValueError(
            f"No MTD sheet found in template.\nSheets: {names}\n"
            "Set MTD_SHEET_NAME at the top of variance_engine.py.")
    if ytd is None:
        raise ValueError(
            f"No YTD sheet found in template.\nSheets: {names}\n"
            "Set YTD_SHEET_NAME at the top of variance_engine.py.")
    return mtd, ytd


# ═══════════════════════════════════════════════════════════════
#  AUDIT LOG
# ═══════════════════════════════════════════════════════════════

def _write_input_log(ws, args: dict, pairs: list):
    ws["A1"] = "INPUT FILE LOG"
    ws["A1"].font = Font(name="Helvetica", bold=True, size=12, color="C3002F")

    meta = [
        ("Generated",   datetime.now().strftime("%d %b %Y  %H:%M")),
        ("Quarter",     args.get("quarter", "")),
        ("Month",       args.get("month", "")),
        ("Master File", args.get("master_file", "")),
        ("Template",    args.get("template_file", "")),
    ]
    lf = Font(name="Helvetica", size=9, bold=True, color="888888")
    vf = Font(name="Helvetica", size=9,             color="C0C0C0")
    for i, (label, value) in enumerate(meta, start=2):
        ws.cell(row=i, column=1, value=label).font = lf
        ws.cell(row=i, column=2, value=value).font = vf

    hf    = PatternFill("solid", fgColor="1C1C1C")
    hfont = Font(name="Helvetica", bold=True, color="C0C0C0", size=9)

    row = len(meta) + 3
    ws.cell(row=row, column=1, value="SCENARIO PAIRS").font = Font(
        name="Helvetica", bold=True, size=9, color="C3002F")
    row += 1
    for col, h in enumerate(
            ["#", "Scenario 1", "Scenario 2", "MTD Tab", "YTD Tab"], 1):
        c = ws.cell(row=row, column=col, value=h)
        c.fill = hf
        c.font = hfont
    row += 1
    for idx, (sc1, sc2) in enumerate(pairs, 1):
        lbl = f"{sc1} vs {sc2}"
        ws.cell(row=row, column=1, value=idx).font               = Font(size=8, color="666666")
        ws.cell(row=row, column=2, value=sc1).font               = Font(size=8, color="C0C0C0")
        ws.cell(row=row, column=3, value=sc2).font               = Font(size=8, color="C0C0C0")
        ws.cell(row=row, column=4, value=f"MTD {lbl}"[:31]).font = Font(size=8, color="888888")
        ws.cell(row=row, column=5, value=f"YTD {lbl}"[:31]).font = Font(size=8, color="888888")
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="INPUT FILES").font = Font(
        name="Helvetica", bold=True, size=9, color="C3002F")
    row += 1
    for col, h in enumerate(["#", "Folder", "File Name", "Full Path"], 1):
        c = ws.cell(row=row, column=col, value=h)
        c.fill = hf
        c.font = hfont
    row += 1
    for i, fp in enumerate(args.get("input_files", []), 1):
        ws.cell(row=row, column=1, value=i).font                    = Font(size=8, color="666666")
        ws.cell(row=row, column=2, value=os.path.dirname(fp)).font  = Font(size=8, color="888888")
        ws.cell(row=row, column=3, value=os.path.basename(fp)).font = Font(size=8, color="C0C0C0")
        ws.cell(row=row, column=4, value=fp).font                   = Font(size=8, color="444444")
        row += 1

    for col, width in zip("ABCDE", [6, 20, 20, 30, 45]):
        ws.column_dimensions[col].width = width
    ws.sheet_view.showGridLines = False


# ═══════════════════════════════════════════════════════════════
#  STANDALONE TEST
# ═══════════════════════════════════════════════════════════════

if __name__ == "__main__":
    from datetime import datetime as _dt
    test_args = {
        "scenario_pairs":  [("FC 2+10", "Actuals"), ("Budget", "LY Actuals")],
        "quarter":         "Q1 (Apr-Jun)",
        "month":           "April",
        "input_folders":   [r"C:\Test\InputData"],
        "input_files":     [r"C:\Test\InputData\Sales_Data.xlsx"],
        "master_file":     r"C:\Test\InputData\Variance_Master.xlsx",
        "template_file":   r"C:\Test\Templates\Variance_Template.xlsx",
        "output_folder":   r"C:\Test\Output",
        "timestamp":       _dt.now().strftime("%Y%m%d_%H%M%S"),
    }
    out = run_variance(test_args)
    print(f"Saved to: {out}")
