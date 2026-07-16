
 
 
 
 
 
import pandas as pd
import openpyxl
from datetime import datetime
import math
import calendar
import os
import shutil

from link_repair import repair_external_links

# ── read-only workbook cache (kills repeated re-parsing of the same template) ──
_WB_READ_CACHE: dict = {}


def _cached_workbook(file_path):
    st = os.stat(file_path)
    key = (os.path.abspath(file_path), st.st_mtime_ns, st.st_size)
    wb = _WB_READ_CACHE.get(key)
    if wb is None:
        wb = openpyxl.load_workbook(file_path, data_only=False, keep_links=True)
        _WB_READ_CACHE.clear()
        _WB_READ_CACHE[key] = wb
    return wb


global column_indices, goodwill_rows, ghq_cfmi_rows, nmgb_buyback_rows, ger_tcs_rows, act_rows, mobility_rows
current_year = datetime.now().year
 
current_date = datetime.now()
 
if current_date.month >= 4:  # If the current month is April or later
    financial_year_start = current_date.year
    financial_year_end = current_date.year + 1
else:  # If the current month is before April
    financial_year_start = current_date.year - 1
    financial_year_end = current_date.year
 
current_year = f"{financial_year_start}-{str(financial_year_end)[-2:]}"
 
def find_row_numbers_act(df):
    # Convert search terms to lowercase and remove spaces for case-insensitive matching
    search_terms = [
        "nissan retail volume", "infiniti retail volume", "datsun retail volume",
        "registration nissan", "registration infiniti", "registration datsun",
        "tiv", "market share", "european wholesale volume - nissan",
        "european wholesale volume - infiniti", "european wholesale volume - datsun",
        "market share - nissan", "market share - infiniti", "market share - datsun",
        "opening stock", "nmuk", "nmisa", "nmgr", "production volume",
        "purchase volume", "total in", "export volume nissan",
        "export volume - datsun", "export volume - infiniti", "export volume",
        "wholesale volume total", "renault trucks", "daimler", "renault",
        "gtr [idg]", "oem volume", "captives", "total out",
        "scope adjustment", "closing stock"
    ]
    search_terms = [term.replace(" ", "").lower() for term in search_terms]
    
    # Convert all data in the DataFrame to lowercase strings and remove spaces for case-insensitive searching
    df_lower = df.apply(lambda x: x.astype(str).str.lower().str.replace(" ", ""))
 
    # Initialize a dictionary to store row indices for each search term
    row_indices = {term: df_lower[df_lower.apply(lambda row: term in row.values, axis=1)].index.tolist() for term in search_terms}
 
    return row_indices
 
class SearchResults:
    wholesale_volume = None
    export_volume = None
    gross_sales_wo_rd = None
    rd_revenues = None
    ci = None
    net_sales = None
    parts_materials_and_others = None
    rd_costs = None
    mfg_fixed = None
    mfg_fixed_plants = None
    mfg_fixed_out_of_scope = None
    vendor_tooling = None
    cogs = None
    cogs_net_sales_percent = None
    rmp_i = None
    rmp_i_percent = None
    gross_profit = None
    fmi = None
    warranty_rs = None
    ga_and_other_items = None
    sga_total = None
    rop = None
    below_op = None
    net_income = None
 
class Goodwill_rows:
    def __init__(self):
        self.column_indices = []
        self.goodwill_rows = []
        self.ghq_cfmi_rows = []
        self.nmgb_buyback_rows = []
        self.ger_tcs_rows = []
        self.act_rows = []
        self.mobility_rows = []
 
# Creating an instance and appending to its list
instance = Goodwill_rows()
 
def find_values_pandas_HFM(df):
    # Convert the DataFrame to lowercase and strip extra spaces
    search_df = df.iloc[0:34].applymap(lambda x: str(x).strip().lower())
 
    # Define search terms with stripped spaces and lowercase conversion
    search_terms = {
        'wholesale volume': 'wholesale_volume',
        'export volume': 'export_volume',
        'gross sales (w/o r&d)': 'gross_sales_wo_rd',
        'r&d revenues': 'rd_revenues',
        'c&i': 'ci',
        'net sales': 'net_sales',
        'parts/materials and others': 'parts_materials_and_others',
        'r&d costs': 'rd_costs',
        'mfg fixed': 'mfg_fixed',
        'mfg fixed - plants': 'mfg_fixed_plants',
        'mfg fixed - out of scope': 'mfg_fixed_out_of_scope',
        'vendor tooling': 'vendor_tooling',
        'cogs': 'cogs',
        '% ( = cogs / net sales)': 'cogs_net_sales_percent',
        'rmp-i': 'rmp_i',
        '% rmp-i': 'rmp_i_percent',
        'gross profit': 'gross_profit',
        'fmi': 'fmi',
        'warranty/rs': 'warranty_rs',
        'g&a and other items': 'ga_and_other_items',
        'sg&a total': 'sga_total',
        'rop': 'rop',
        'below op': 'below_op',
        'net income': 'net_income'
    }
 
    # Find rows matching each search term and store in static variables
    for term, attr in search_terms.items():
        row_index = search_df[search_df.apply(lambda row: row.str.contains(term, regex=False).any(), axis=1)].index
        setattr(SearchResults, attr, int(row_index[0]) if not row_index.empty else None)
 
def generate_excel_formula(row_num, col_num, sheet_name=None, file_name=None):
    excel_row = row_num + 2
    excel_column = col_num + 1
    def column_letter(col_num):
        letter = ''
        while col_num > 0:
            col_num, remainder = divmod(col_num - 1, 26)
            letter = chr(65 + remainder) + letter
        return letter
    
    excel_col_letter = column_letter(excel_column)
 
    if file_name and sheet_name:
        corrected_file_name = file_name.replace("\\", "/")
        formula = f"='[{corrected_file_name}]{sheet_name}'!{excel_col_letter}{excel_row}"
    elif sheet_name:
        formula = f"='{sheet_name}'!{excel_col_letter}{excel_row}"
    else:
        formula = f"={excel_col_letter}{excel_row}"
    
    return formula
 
def generate_excel_formula_1(row_num, col_num, sheet_name=None, file_name=None):
    excel_row = row_num 
    excel_column = col_num + 1
    def column_letter(col_num):
        letter = ''
        while col_num > 0:
            col_num, remainder = divmod(col_num - 1, 26)
            letter = chr(65 + remainder) + letter
        return letter
    
    excel_col_letter = column_letter(excel_column)
 
    if file_name and sheet_name:
        corrected_file_name = file_name.replace("\\", "/")
        formula = f"='[{corrected_file_name}]{sheet_name}'!{excel_col_letter}{excel_row}"
    elif sheet_name:
        formula = f"='{sheet_name}'!{excel_col_letter}{excel_row}"
    else:
        formula = f"={excel_col_letter}{excel_row}"
    
    return formula    
 
def generate_excel_formula_2(row_num, col_num, sheet_name=None, file_name=None):
    excel_row = row_num+2 
    excel_column = col_num
    def column_letter(col_num):
        letter = ''
        while col_num > 0:
            col_num, remainder = divmod(col_num - 1, 26)
            letter = chr(65 + remainder) + letter
        return letter
    
    excel_col_letter = column_letter(excel_column)
 
    if file_name and sheet_name:
        corrected_file_name = file_name.replace("\\", "/")
        formula = f"='[{corrected_file_name}]{sheet_name}'!{excel_col_letter}{excel_row}"
    elif sheet_name:
        formula = f"='{sheet_name}'!{excel_col_letter}{excel_row}"
    else:
        formula = f"={excel_col_letter}{excel_row}"
    
    return formula    
 
def extract_formula(file_path, sheet_name, row_number, column_index):
        try:
            # Load workbook (cached, read-only, keep_links) and select sheet
            workbook = _cached_workbook(file_path)
            sheet = workbook[sheet_name]
 
            # Access the cell using row and column index
            cell = sheet.cell(row=row_number, column=column_index + 1)  # openpyxl is 1-indexed
 
            # Extract and print the formula
            formula = cell.value if isinstance(cell.value, str) and cell.value.startswith('=') else None
            print(f"Formula at row {row_number}, column index {column_index}: {formula}")
 
            return formula
        except KeyError:
            print("Invalid sheet name provided.")
        except IndexError:
            print("Invalid row or column index provided.")
 
def paste_values_KPI_PL(original_file_path, sheet_names, cell_updates_list, month_name,output_path):
    # keep_links=True preserves the input-file links already present in the template
    workbook = openpyxl.load_workbook(original_file_path, keep_links=True)

    for sheet_name, cell_updates in zip(sheet_names, cell_updates_list):
        sheet = workbook[sheet_name]

        for update in cell_updates:
            row, column, formula_or_string = update
            cell = sheet.cell(row=row, column=column)
            if formula_or_string is not None:
                cell.value = formula_or_string

    month_name = calendar.month_name[month_name]
    os.makedirs(output_path, exist_ok=True)
    reusult = os.path.join(output_path, f'NEW_KPI_P&L_{month_name}_output_FY{current_year}.xlsx')

    # Save ONCE to the output folder (dead 'New KPI Files' / 'Mandatory_folder'
    # duplicates removed – they were never read back).
    workbook.save(reusult)
    workbook.close()

    # Register cross-workbook links so Excel opens without the repair dialog.
    try:
        repair_external_links(reusult)
    except Exception as e:
        print(f"[WARN] link repair skipped for {reusult}: {e}")

    print(reusult)
    return reusult
 
# def paste_values_KPI_PL(original_file_path, sheet_names, cell_updates_list, month_number):
 
#     # Load workbook
#     workbook = openpyxl.load_workbook(original_file_path, data_only=False)
 
#     # Apply updates to each sheet
#     for sheet_name, cell_updates in zip(sheet_names, cell_updates_list):
#         sheet = workbook[sheet_name]
#         for row, column, formula_or_string in cell_updates:
#             if formula_or_string is not None:
#                 cell = sheet.cell(row=row, column=column)
#                 cell.value = formula_or_string
 
#     # Prepare output directories
#     output_dir_1 = os.path.join(os.getcwd(), 'New KPI Files')
#     output_dir_2 = os.path.join(os.getcwd(), 'Mandatory_folder')
#     os.makedirs(output_dir_1, exist_ok=True)
#     os.makedirs(output_dir_2, exist_ok=True)
 
#     # Format filename
#     month_name = calendar.month_name[month_number]
#     current_year = datetime.now().year
#     file_name = f'NEW_KPI_P&L_{month_name}_output_FY{current_year}.xlsx'
 
#     new_file_path = os.path.join(output_dir_1, file_name)
#     new_file_path_1 = os.path.join(output_dir_2, file_name)
 
#     # Save workbook to both locations
#     workbook.save(new_file_path)
#     workbook.save(new_file_path_1)
 
#     print(f"File saved successfully at:\n- {new_file_path}\n- {new_file_path_1}")
#     return new_file_path
 
def relief_search_terms_new(df):
    # Define the search terms with proper formatting
    search_terms = [
        "recall service/campaigns",
        "kd itp (nmuk)",
        "kd itp (nmisa)",
        "kd itp (nmgr)",
        "connectivity",
        "itp cbu nissan",
        "itp cbu infiniti",
        "iln export (nmuk)",
        "iln export (nmisa)",
        "error correction in mz (?)",
        "export",
        "mz impairment",
        "g&a impairment",
        "compact ev/vt impairment",
        "battery 40 kw itp ( nmuk )",
        "battery 40 kw itp ( nmisa )",
        "battery cancelation",
        "error correction in mz (after wd5)",
        "top daimler correction (after wd5)",
        "infiniti fmi correction (after wd5)",
        "G&A Impairment",
        "itp cbu nissan",
        "itp cbu infiniti",
        "connectivity (ccs)"
    ]
 
    # Preprocess the dataframe: convert to lowercase and remove extra spaces
    search_df = df.iloc[70:].apply(lambda x: x.astype(str).str.lower().str.replace(r'\s+', '', regex=True))
    results = {}
 
    # Preprocess search terms to remove spaces and handle special characters
    processed_terms = [term.replace(' ', '').replace('(', r'\(').replace(')', r'\)').replace('?', r'\?') for term in search_terms]
 
    # Iterate over each search term to find its occurrence in the dataframe
    for term, processed_term in zip(search_terms, processed_terms):
        # Find the index of the first occurrence of the term in any row
        row_index = search_df[search_df.apply(lambda row: row.str.contains(processed_term, regex=True).any(), axis=1)].index
        
        # Store the result in the dictionary, indicating 'Not found' if no occurrence is found
        results[term] = row_index[0] if not row_index.empty else 'Not found'
 
    # # Print the results for each term
    # for term, index in results.items():
    #     print(f"'{term}' found at row: {index}")
    
    return results 
 
def find_row_number_CMFI(df, search_value='NML Cemtral FMI', start_row=0, end_row=30):
 
    row_numbers = []
    for index, row in df.iterrows():
        if start_row <= index <= end_row:
            if search_value in row.values:
                row_numbers.append(index)
 
    return row_numbers   
def find_row_and_column(df, column_number):
    # Define the month mapping ignoring the year, starting from April
    month_mapping = {
        4: "Apr",
        5: "May",
        6: "Jun",
        7: "Jul",
        8: "Aug",
        9: "Sep",
        10: "Oct",
        11: "Nov",
        12: "Dec",
        1: "Jan",
        2: "Feb",
        3: "Mar"
    }
    
    # Validate the column number
    if column_number not in month_mapping:
        raise ValueError("Invalid column number. Please provide a number between 1 and 12.")
    
    # Convert the entire DataFrame to lowercase for case-insensitive search
    df_lower = df.apply(lambda x: x.astype(str).str.lower())
    
    # Search for the term "gw after scope change" in all columns
    search_term = "gw after scope change"
    occurrences = df_lower.apply(lambda row: row.str.contains(search_term, na=False)).any(axis=1)
    
    # Find all row indices where the search term occurs
    row_indices = occurrences[occurrences].index.tolist()
    
    if len(row_indices) < 2:
        raise ValueError("The term 'GW after scope change' was not found twice in any column.")
    
    # Return the index of the second occurrence
    return row_indices[1], column_number
 
def OSP_CALc(df, month_number):
    # Adjust the month index to start from April
    adjusted_month_index = (month_number - 4) % 12
    
    # Define the search term
    search_term = "outsideprofit"
    
    for row_index, row in df.iterrows():
        for col_index, value in enumerate(row):
            # Clean the value: remove spaces and convert to lowercase
            cleaned_value = str(value).replace(" ", "").lower()
            
            if cleaned_value == search_term:
                # Calculate the target column index based on the adjusted month index
                target_col_index = col_index + adjusted_month_index
                month_value = df.iloc[0, target_col_index]
                
                return row_index + 1, col_index + adjusted_month_index
    
    return None, None        
 
def find_itp_and_month_value(df, month):
    adjusted_month_index = (month - 4) % 12
    
    for row_index, row in df.iterrows():
        for col_index, value in enumerate(row):
            if value == "ITP":
                target_col_index = col_index + adjusted_month_index
                month_value = df.iloc[0, target_col_index]
                return row_index -2, target_col_index
    
    return "Not Found", None
 
def find_second_column_for_outside_profit(df,month_number):
    # Define the search term
    search_term = "outsideprofit"
    adjusted_month_index = (month_number - 4) % 12
    
    # Initialize a counter to track occurrences
    occurrence_count = 0
    
    for row_index, row in df.iterrows():
        for col_index, value in enumerate(row):
            # Clean the value: remove spaces and convert to lowercase
            cleaned_value = str(value).replace(" ", "").lower()
            
            if cleaned_value == search_term:
                occurrence_count += 1
                if occurrence_count == 2:
                    return row_index - 2, col_index + adjusted_month_index
    
    return None, None
 
def goodwill_raw_data_extraction(df, month_number, search_terms):
    # Convert DataFrame to lowercase and remove spaces for case-insensitive comparison
    df_lower = df.apply(lambda x: x.astype(str).str.lower().str.replace(" ", "") if x.dtype == 'object' else x)
    
    # Initialize lists to store indices
    column_indices = []
    goodwill_rows = []
    ghq_cfmi_rows = []
    nmgb_buyback_rows = []
    ger_tcs_rows = []
    act_rows = []
    mobility_rows = []
 
    # Find columns with dates matching the given month number
    for col in df.columns:
        if df[col].dtype == object:  # Check if column contains strings (dates)
            for idx, date_str in enumerate(df[col]):
                try:
                    # Ensure date_str is a string before parsing
                    date_obj = datetime.strptime(str(date_str), '%m/%d/%Y')
                    if date_obj.month == month_number:
                        column_indices.append(col)
                        break
                except ValueError:
                    continue
    
    # Find rows matching the search terms exactly
    for idx, row in df_lower.iterrows():
        for term in search_terms:
            term_cleaned = term.lower().replace(" ", "")
            if any(term_cleaned == value for value in row):
                if term_cleaned == "goodwill":
                    goodwill_rows.append(idx)
                elif term_cleaned == "ghq(cfmi)":
                    ghq_cfmi_rows.append(idx)
                elif term_cleaned == "nmgbbuyback-scopechangetohq":
                    nmgb_buyback_rows.append(idx)
                elif term_cleaned == "gertcs":
                    ger_tcs_rows.append(idx)
                elif term_cleaned == "act":
                    act_rows.append(idx)
                elif term_cleaned == "mobility":
                    mobility_rows.append(idx)        
                break
    
    # Return all variables separately
    return column_indices, goodwill_rows, ghq_cfmi_rows, nmgb_buyback_rows, ger_tcs_rows, act_rows, mobility_rows



# def paste_values_goodwill(original_file_path, sheet_names, cell_updates_list, month_name,output_path):
#     workbook = openpyxl.load_workbook(original_file_path,data_only=False)
 
#     for sheet_name, cell_updates in zip(sheet_names, cell_updates_list):
#         sheet = workbook[sheet_name]
 
#         for update in cell_updates:
#             row, column, formula_or_string = update
#             cell = sheet.cell(row=row, column=column)
#             if formula_or_string is not None:
#                 if isinstance(formula_or_string, str) and formula_or_string.startswith('='):
#                     cell.value = formula_or_string
#                 else:
#                     cell.value = formula_or_string
 
#     timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
 
#     month_name = calendar.month_name[month_name]
 
#     new_file_path = os.path.join(os.getcwd(), 'Goodwill_files', f'Goodwill_{month_name}_output_FY{current_year}.xlsx')
#     reusult=output_path+"\\"+f'Goodwill_{month_name}_output_FY{current_year}.xlsx'
 
#     workbook.save(new_file_path)
#     workbook.close()
 
#     workbook.save(reusult)
 
def paste_values_goodwill(original_file_path, sheet_names, cell_updates_list, month_name,output_path):
    # keep_links=True preserves existing input-file links in the goodwill template
    workbook = openpyxl.load_workbook(original_file_path, keep_links=True)
    for sheet_name, cell_updates in zip(sheet_names, cell_updates_list):
        sheet = workbook[sheet_name]
        for update in cell_updates:
            row, column, formula_or_string = update
            if formula_or_string is not None:
                sheet.cell(row=row,column=column).value=formula_or_string

    month_str=calendar.month_name[month_name]
    file_name=f"Goodwill_{month_str}_output_FY{current_year}.xlsx"

    result=os.path.join(output_path,file_name)
    os.makedirs(os.path.dirname(result),exist_ok=True)

    # Save ONCE, then repair links on that single file.
    workbook.save(result)
    workbook.close()
    try:
        repair_external_links(result)
    except Exception as e:
        print(f"[WARN] link repair skipped for {result}: {e}")

    # A copy in Goodwill_files/ is REQUIRED: it is read back as next month's
    # goodwill template.  Copy the already-repaired file byte-for-byte (fast,
    # link-clean) instead of re-serialising the workbook again.
    new_file_path = os.path.join(os.getcwd(), "Goodwill_files", file_name)
    os.makedirs(os.path.dirname(new_file_path), exist_ok=True)
    shutil.copy2(result, new_file_path)
       
 
# def paste_values_goodwill(original_file_path, sheet_names, cell_updates_list, month_number):
#     # Load workbook (preserving formulas)
#     workbook = openpyxl.load_workbook(original_file_path, data_only=False)
 
#     # Apply updates to each sheet
#     for sheet_name, cell_updates in zip(sheet_names, cell_updates_list):
#         sheet = workbook[sheet_name]
#         for row, column, formula_or_string in cell_updates:
#             if formula_or_string is not None:
#                 cell = sheet.cell(row=row, column=column)
#                 cell.value = formula_or_string
 
#     # Prepare output directory
#     output_dir = os.path.join(os.getcwd(), 'Goodwill_files')
#     os.makedirs(output_dir, exist_ok=True)
 
#     # Format filename
#     month_name = calendar.month_name[month_number]
#     current_year = datetime.now().year
#     new_file_name = f'Goodwill_{month_name}_output_FY{current_year}.xlsx'
#     new_file_path = os.path.join(output_dir, new_file_name)
 
#     # Save workbook once
#     workbook.save(new_file_path)
 
#     # print(f"File saved successfully at: {new_file_path}")
 




