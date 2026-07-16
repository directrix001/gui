 
"""
Submit_2_Helper_function.py
============================
All row/column search helpers, formula generators, and the main
paste_values_KPI_PL writer used by Submit_2.py.
 
KEY FIX – Excel link preservation
----------------------------------
openpyxl strips external workbook relationships when it loads and
re-saves a file unless you explicitly tell it to keep them.
Every load_workbook() call that writes back uses:
    keep_links=True   – preserves [xl/externalLinks] relationships
    keep_vba=False    – we don't need macros (set True if the template has VBA)
    data_only=False   – we WRITE formulas, so we never want data_only=True here
 
Cross-workbook formula strings ("='[C:/path/file.xlsx]Sheet'!A1") must use
forward-slashes and must NOT be modified after they are built by
generate_excel_formula(); openpyxl writes them verbatim.
"""
 
from __future__ import annotations
 
import calendar
import os
import re
from datetime import datetime
from typing import Optional
 
import numpy as np
import openpyxl
import pandas as pd

from link_repair import repair_external_links

# ─── fiscal-year label used in output file names ───────────────────────────
_now = datetime.now()
_fy_start = _now.year if _now.month >= 4 else _now.year - 1
CURRENT_FY = f"{_fy_start}-{str(_fy_start + 1)[-2:]}"


# ══════════════════════════════════════════════════════════════════════════════
#  Read-only workbook cache
#  ------------------------------------------------------------------------------
#  extract_formula() is called dozens of times per run, almost always against
#  the SAME template file.  Re-parsing a large .xlsx on every call was the main
#  time sink.  We cache the loaded (read-only) workbook keyed by (path, mtime,
#  size) so a modified file is automatically re-read.  These handles are only
#  ever used for READS, never saved, so the cache is safe for the whole run.
# ══════════════════════════════════════════════════════════════════════════════
_WB_READ_CACHE: dict = {}


def _cached_workbook(file_path: str):
    st = os.stat(file_path)
    key = (os.path.abspath(file_path), st.st_mtime_ns, st.st_size)
    wb = _WB_READ_CACHE.get(key)
    if wb is None:
        wb = openpyxl.load_workbook(file_path, data_only=False, keep_links=True)
        _WB_READ_CACHE.clear()          # keep only the most-recent file (low memory)
        _WB_READ_CACHE[key] = wb
    return wb


def clear_workbook_cache() -> None:
    for wb in _WB_READ_CACHE.values():
        try:
            wb.close()
        except Exception:
            pass
    _WB_READ_CACHE.clear()
 
 
# ══════════════════════════════════════════════════════════════════════════════
#  Section 1 – Row / column search helpers
# ══════════════════════════════════════════════════════════════════════════════
 
def find_row_numbers(df: pd.DataFrame) -> dict[str, list[int]]:
    """
    Return a dict mapping normalised search-term → list of matching row indices
    found anywhere in *df*.  Comparison is space-stripped, lowercase.
    """
    search_terms = [
        "nissan retail volume", "infiniti retail volume", "datsun retail volume",
        "registration nissan", "registration infiniti", "registration datsun",
        "tiv", "market share",
        "european wholesale volume - nissan",
        "european wholesale volume - infiniti",
        "european wholesale volume - datsun",
        "market share - nissan", "market share - infiniti", "market share - datsun",
        "opening stock", "nmuk", "nmisa", "nmgr", "production volume",
        "purchase volume", "total in",
        "export volume nissan", "export volume - datsun", "export volume - infiniti",
        "export volume", "wholesale volume total",
        "renault trucks", "daimler", "renault",
        "gtr [idg]", "oem volume", "captives", "total out",
        "scope adjustment", "closing stock",
    ]
    normalised = [t.replace(" ", "").lower() for t in search_terms]
    df_norm = df.apply(lambda col: col.astype(str).str.lower().str.replace(" ", "", regex=False))
    return {
        term: df_norm[df_norm.apply(lambda row: term in row.values, axis=1)].index.tolist()
        for term in normalised
    }
 
 
def search_turnover_in_dataframe(df: pd.DataFrame) -> dict[str, list[int]]:
    """
    Find 'turnover' in the first 4 columns then search ±22 rows for
    brand-level sub-category rows.
    """
    df_lower = df.apply(lambda col: col.astype(str).str.lower())
    turnover_row: Optional[int] = None
    for col in df_lower.columns[:4]:
        hits = df_lower[df_lower[col].str.contains("turnover", na=False)].index
        if not hits.empty:
            turnover_row = hits[0]
            break
    if turnover_row is None:
        return {}
 
    search_terms = [
        "nissan nv", "nissan bb", "datsun nv", "datsun bb",
        "infiniti nv", "infiniti bb",
        "infiniti used cars", "nissan used cars",
        "nissan after sales", "datsun after sales", "infiniti after sales",
        "export nissan", "export datsun", "export infiniti",
    ]
    result: dict[str, list[int]] = {t: [] for t in search_terms}
    window = df_lower.iloc[turnover_row: turnover_row + 22]
    for term in search_terms:
        for col in window.columns:
            result[term].extend(
                window[window[col].str.contains(term, na=False)].index.tolist()
            )
    return result
 
 
def _find_fmi_indices(df: pd.DataFrame, start: int) -> Optional[tuple[int, int, int, int]]:
    """Within df[start : start+8] find rows for FMI, TurnoverAj, COS Regional, VME."""
    end = start + 8
    block = df.iloc[start:end]
    fmi   = block[block.isin(["FMI"]).any(axis=1)].index
    trn   = block[block.isin(["TurnoverAj"]).any(axis=1)].index
    cos_  = block[block.isin(["COS Regional"]).any(axis=1)].index
    vme   = block[block.isin(["VME"]).any(axis=1)].index
    if all(not x.empty for x in (fmi, trn, cos_, vme)):
        return fmi[0], trn[0], cos_[0], vme[0]
    return None
 
 
def _brand_starts(df: pd.DataFrame, *labels: str) -> list[int]:
    """Return first-occurrence row index for each label (lowercase, first 3 cols)."""
    df_lower = df.iloc[:, 0:3].apply(
        lambda col: col.str.lower() if col.dtype == "object" else col
    )
    result = []
    for label in labels:
        idx = df_lower[df_lower.isin([label]).any(axis=1)].index
        if idx.empty:
            raise ValueError(f"Required label '{label}' not found in first 3 columns.")
        result.append(idx[0])
    return result
 
 
# def find_fmi_row_nv(df: pd.DataFrame):
#     starts = _brand_starts(df, "nissan nv", "datsun nv", "infiniti nv", "total nv")
#     rows = [_find_fmi_indices(df, s) for s in starts]
#     if any(r is None for r in rows):
#         raise ValueError("FMI indices incomplete for NV section.")
#     return tuple(v for r in rows for v in r)
def find_fmi_row_nv(df: pd.DataFrame, fallback_rows: tuple | None = None):
    starts = _brand_starts(df, "nissan nv", "datsun nv", "infiniti nv", "total nv")
    rows = [_find_fmi_indices(df, s) for s in starts]
    
    if any(r is None for r in rows):
        if fallback_rows is not None:
            return fallback_rows
        raise ValueError("FMI indices incomplete for NV section.")
    
    return tuple(v for r in rows for v in r)

 


 
 
def find_fmi_row_uc(df: pd.DataFrame):
    starts = _brand_starts(df, "uc nissan", "uc infiniti", "total uc")
    rows = [_find_fmi_indices(df, s) for s in starts]
    if any(r is None for r in rows):
        raise ValueError("FMI indices incomplete for UC section.")
    return tuple(v for r in rows for v in r)
 
 
def find_fmi_row_as(df: pd.DataFrame):
    starts = _brand_starts(df, "as nissan", "as datsun", "as infiniti", "total as")
    rows = [_find_fmi_indices(df, s) for s in starts]
    if any(r is None for r in rows):
        raise ValueError("FMI indices incomplete for AS section.")
    return tuple(v for r in rows for v in r)
 
 
def find_fmi_row_export(df: pd.DataFrame):
    starts = _brand_starts(df,
                           "nissan export", "datsun export", "infiniti export", "total export")
    rows = [_find_fmi_indices(df, s) for s in starts]
    if any(r is None for r in rows):
        raise ValueError("FMI indices incomplete for Export section.")
    return tuple(v for r in rows for v in r)
 
 
def _search_terms_from_anchor(
    df: pd.DataFrame,
    anchor_substring: str,
    search_terms: list[str],
    window: int = 28,
) -> dict[str, list[int]]:
    """Generic: find anchor in first 4 cols, then search *window* rows below."""
    df_lower = df.apply(lambda col: col.astype(str).str.lower())
    anchor_row: Optional[int] = None
    for col in df_lower.columns[:4]:
        hits = df_lower[df_lower[col].str.contains(anchor_substring, na=False)].index
        if not hits.empty:
            anchor_row = hits[0]
            break
    if anchor_row is None:
        return {t: [] for t in search_terms}
    block = df_lower.iloc[anchor_row: anchor_row + window]
    result: dict[str, list[int]] = {t: [] for t in search_terms}
    for term in search_terms:
        pat = re.escape(term)
        for col in block.columns:
            result[term].extend(
                block[block[col].str.contains(pat, na=False)].index.tolist()
            )
    return result
 
 
def search_omp_terms(df: pd.DataFrame) -> dict[str, list[int]]:
    return _search_terms_from_anchor(df, "omp", [
        "x83 sales to renault",
        "f91r sales to renault trucks",
        "production parts sales to inter-companies",
        "cantabria sales to third party",
        "parts sales to renault",
        "battery plant other costs",
        "scraps sales",
        "pensions",
        "others",
    ])
 
 
def search_manufacturing_terms(df: pd.DataFrame) -> dict[str, list[int]]:
    return _search_terms_from_anchor(df, "manufacturing", [
        "materials costs",
        "b/o cost reduction",
        "renault tp",
        "intercompany parts / kd",
        "inbound freight / insurance",
        "inbound duty",
        "inhouse / direct labour",
        "battery (nmuk)",
        "events",
        "infiniti mz variable",
        "other adjustments",
        "others",
        "core fixed",
        "pic (nissan)",
        "regional cost",
        "vendor tooling depreciation",
        "projects",
        "nesas/nisa manufacturing",
        "battery fixed",
        "infiniti fixed cost (gd1a)",
        "maintenance of existing models (d&d)",
        "gd1a top position shift to icp",
        "supplier claims & other",
    ])
 
 
def MZ_Sheet_terms(df: pd.DataFrame) -> dict[str, list[int]]:
    return _search_terms_from_anchor(df, "turnover oem/ omp", [
        "x83 sales to renault",
        "sales to renault trucks",
        "production parts sales to inter-companies",
        "cantabria sales to third party",
        "parts sales to renault (h4bu camshaft)",
        "scrap sales",
        "battery plant to",
        "others",
        "other adjustments",
    ], window=10)
 
 
def find_financial_indices(df: pd.DataFrame):
    """Return (labour, IS, dep, provision, overhead) index lists from G&A block."""
    df_lower = df.iloc[:, 0:3].apply(
        lambda col: col.str.lower() if col.dtype == "object" else col
    )
    starts = df_lower[df_lower.isin(["g&a"]).any(axis=1)].index
    if starts.empty:
        raise ValueError("'g&a' not found in first 3 columns.")
    start = starts[0]
    end   = min(start + 24, len(df))
    block = df.iloc[start:end]
    return (
        block[block.isin(["Labour Costs"]).any(axis=1)].index,
        block[block.isin(["IS Costs"]).any(axis=1)].index,
        block[block.isin(["Depreciation"]).any(axis=1)].index,
        block[block.isin(["Provision for Doubtful Debts"]).any(axis=1)].index,
        block[block.isin(["Overheads"]).any(axis=1)].index,
    )
 
 
def find_oem_terms(df: pd.DataFrame):
    df_lower = df.iloc[:, 0:3].apply(
        lambda col: col.str.lower() if col.dtype == "object" else col
    )
    starts = df_lower[df_lower.isin(["oem"]).any(axis=1)].index
    if starts.empty:
        raise ValueError("'oem' not found in first 3 columns.")
    start = starts[0]
    end   = min(start + 24, len(df))
    block = df.iloc[start:end]
    return (
        block[block.isin(["Turnover"]).any(axis=1)].index,
        block[block.isin(["Standard Cost"]).any(axis=1)].index,
        block[block.isin(["Marginal Profit"]).any(axis=1)].index,
        block[block.isin(["Vehicle enhancement"]).any(axis=1)].index,
        block[block.isin(["Logistics"]).any(axis=1)].index,
        block[block.isin(["Royalty"]).any(axis=1)].index,
        block[block.isin(["Others"]).any(axis=1)].index,
        block[block.isin(["EOP"]).any(axis=1)].index,
    )
 
 
def find_ocs_terms(df: pd.DataFrame):
    df_lower = df.iloc[:, 0:3].apply(
        lambda col: col.str.lower() if col.dtype == "object" else col
    )
    starts = df_lower[df_lower.isin(["ocs"]).any(axis=1)].index
    if starts.empty:
        raise ValueError("'ocs' not found in first 3 columns.")
    return starts
 
 
def search_rd_terms(df: pd.DataFrame) -> dict[str, list[int]]:
    df_lower = df.apply(lambda col: col.astype(str).str.lower())
    result: dict[str, list[int]] = {"functional r&d": [], "non functional r&d": []}
    for term in result:
        for col in df_lower.columns:
            result[term].extend(
                df_lower[df_lower[col].str.contains(term, na=False)].index.tolist()
            )
    return result
 
 
def find_R_and_D_File(df: pd.DataFrame) -> dict[str, list[int]]:
    # Use .map() instead of deprecated .applymap()
    df_lower = df.map(lambda x: x.lower() if isinstance(x, str) else x)

    terms = [
        "functional r&d revenue",
        "functional r&d cost",
        "non functional r&d revenue",
        "non functional r&d cost",
    ]

    # Work in positional space entirely — last 15 rows by position
    total_rows = len(df_lower)
    start_pos = max(0, total_rows - 1)          # never go negative
    df_window = df_lower.iloc[start_pos:]        # slice by position

    result = {}
    for t in terms:
        # Build mask only on the window, then convert positional hits to original labels
        mask = df_window.isin([t]).any(axis=1)
        result[t] = df_window.index[mask].tolist()

    return result
 
def find_transfer_price_adjustment(df: pd.DataFrame) -> list[int]:
    pat = re.sub(r"\s+", "", "nrv").lower()
    df_lower = df.applymap(
        lambda x: re.sub(r"\s+", "", x.lower()) if isinstance(x, str) else x
    )
    return df_lower[
        df_lower.apply(lambda row: row.astype(str).str.contains(pat, regex=True).any(), axis=1)
    ].index.tolist()
 
 
def export_ITP_Export(df: pd.DataFrame) -> list[int]:
    pat = re.escape("itp ( revenue )".replace(" ", ""))
    df_lower = df.applymap(
        lambda x: re.sub(r"\s+", "", x.lower()) if isinstance(x, str) else x
    )
    return df_lower[
        df_lower.apply(lambda row: row.astype(str).str.contains(pat, regex=True).any(), axis=1)
    ].index.tolist()
 
 
def energy_tgk_68(df: pd.DataFrame) -> list[int]:
    pat = re.escape("energy business (not inside as for bp)".replace(" ", ""))
    df_lower = df.applymap(
        lambda x: re.sub(r"\s+", "", x.lower()) if isinstance(x, str) else x
    )
    return df_lower[
        df_lower.apply(lambda row: row.astype(str).str.contains(pat, regex=True).any(), axis=1)
    ].index.tolist()
 
 
def tax_and_public(df: pd.DataFrame) -> list[int]:
    pat = "tax and public dues - other tax and public dues"
    df_lower = df.map(lambda x: x.lower() if isinstance(x, str) else x)
    return df_lower[
        df_lower.apply(lambda row: row.astype(str).str.contains(pat).any(), axis=1)
    ].index.tolist()
 
 
def operating_profit(df: pd.DataFrame) -> list[int]:
    pat = "operating profit - aprite"
    df_lower = df.applymap(lambda x: x.lower() if isinstance(x, str) else x)
    return df_lower[
        df_lower.apply(lambda row: row.astype(str).str.contains(pat).any(), axis=1)
    ].index.tolist()
 
 
def relief_search_terms_new(df: pd.DataFrame) -> dict[str, int | str]:
    """
    Search from row 70 onwards for each relief-section label.
    Returns {term: row_index} or {term: 'Not found'}.
    """
    terms = [
        "recall service/campaigns",
        "kd itp (nmuk)", "kd itp (nmisa)", "kd itp (nmgr)",
        "connectivity",
        "itp cbu nissan", "itp cbu infiniti",
        "iln export (nmuk)", "iln export (nmisa)",
        "error correction in mz (?)",
        "export", "mz impairment", "g&a impairment",
        "compact ev/vt impairment",
        "battery 40 kw itp ( nmuk )", "battery 40 kw itp ( nmisa )",
        "battery cancelation",
        "error correction in mz (after wd5)",
        "top daimler correction (after wd5)",
        "infiniti fmi correction (after wd5)",
    ]
    search_df = df.iloc[70:].apply(
        lambda col: col.astype(str).str.lower().str.replace(r"\s+", "", regex=True)
    )
    results: dict[str, int | str] = {}
    for term in terms:
        pat = re.escape(term.replace(" ", "").replace("(", r"\(").replace(")", r"\)").replace("?", r"\?"))
        hits = search_df[
            search_df.apply(lambda row: row.str.contains(pat, regex=True).any(), axis=1)
        ].index
        results[term] = hits[0] if not hits.empty else "Not found"
    return results
 
 
def G_and_A_Reliev1(df: pd.DataFrame) -> Optional[int]:
    """Return first row where any cell starts with 'meur' (case-insensitive)."""
    hits = df[df.apply(
        lambda row: row.astype(str).str.lower().str.startswith("meur").any(), axis=1
    )].index
    return hits[0] if not hits.empty else None
 
 
# ══════════════════════════════════════════════════════════════════════════════
#  Section 2 – Column / month index finders
# ══════════════════════════════════════════════════════════════════════════════
 
def find_columns_with_month_short(df: pd.DataFrame, month_number: int) -> int:
    """
    Find the column index of the cell containing the 3-letter month abbreviation
    (e.g. 'apr') and return the offset column for *month_number*.
    """
    abbr = {4:"apr",5:"may",6:"jun",7:"jul",8:"aug",9:"sep",
            10:"oct",11:"nov",12:"dec",1:"jan",2:"feb",3:"mar"}
    target = abbr[month_number]
    for _, row in df.iterrows():
        for col_name, val in row.items():
            if str(val).lower() == target:
                col_idx = df.columns.get_loc(col_name)
                # offset from April baseline
                return col_idx + (month_number - 4) % 12
    # fallback: calculate from April position
    for _, row in df.iterrows():
        for col_name, val in row.items():
            if str(val).lower() == "apr":
                col_idx = df.columns.get_loc(col_name)
                return col_idx + (month_number - 4) % 12
    return 4 + (month_number - 4) % 12   # hard fallback
 
 
def find_columns_with_month_long(df: pd.DataFrame, month_number: int) -> int:
    """Same as above but looks for the full month name ('april', 'january', …)."""
    full = {1:"january",2:"february",3:"march",4:"april",5:"may",6:"june",
            7:"july",8:"august",9:"september",10:"october",11:"november",12:"december"}
    target = full[month_number]
    for _, row in df.iterrows():
        for col_name, val in row.items():
            if str(val).lower() == target:
                col_idx = df.columns.get_loc(col_name)
                return col_idx + (month_number - 4) % 12
    # fallback: anchor on april
    for _, row in df.iterrows():
        for col_name, val in row.items():
            if str(val).lower() == "april":
                col_idx = df.columns.get_loc(col_name)
                return col_idx + (month_number - 4) % 12
    return 4 + (month_number - 4) % 12
 
 
def find_month_columns_Relief_KD(df: pd.DataFrame, month_number: int) -> list[int]:
    """Return the two column indices where the month name appears in the KD sheet."""
    full = {1:"January",2:"February",3:"March",4:"April",5:"May",6:"June",
            7:"July",8:"August",9:"September",10:"October",11:"November",12:"December"}
    name = full[month_number]
    found: list[int] = []
    for _, row in df.iterrows():
        tmp = [ci for ci, v in enumerate(row)
               if isinstance(v, str) and name in v]
        if len(tmp) >= 2:
            found.extend(tmp[:2])
            break
        elif tmp:
            found.extend(tmp)
        if len(found) >= 2:
            break
    return sorted(set(found))[:2]
 
 
def find_month_columns_export(df: pd.DataFrame, month_number: int) -> list[int]:
    """Return column indices where the ITP month code (M01–M12) appears."""
    codes = {4:"M01",5:"M02",6:"M03",7:"M04",8:"M05",9:"M06",
             10:"M07",11:"M08",12:"M09",1:"M10",2:"M11",3:"M12"}
    code = codes[month_number]
    found: set[int] = set()
    for _, row in df.iterrows():
        for ci, val in enumerate(row):
            if isinstance(val, str) and code in val:
                found.add(ci)
    return sorted(found)
 
 
def find_indices_feuil(df: pd.DataFrame, month_number: int) -> tuple[int, int]:
    """Return (row_index_of_NE_Total, col_index_of_month) in the Feuil1 sheet."""
    df_lower = df.apply(
        lambda row: row.map(lambda x: x.lower() if isinstance(x, str) else x), axis=1
    )
    hits = df_lower[df_lower.apply(lambda row: "ne total" in row.values, axis=1)].index
    if hits.empty:
        raise ValueError("'NE Total' not found in Feuil1.")
    row_idx = hits[0]
    abbr = {1:"jan",2:"feb",3:"mar",4:"apr",5:"may",6:"jun",
            7:"jul",8:"aug",9:"sep",10:"oct",11:"nov",12:"dec"}
    name = abbr[month_number]
    col_idx = df_lower.apply(lambda row: name in row.values, axis=1).idxmax()
    return row_idx, col_idx
 
 
def find_op_and_month(df: pd.DataFrame, month_number: int) -> tuple[Optional[int], Optional[int]]:
    """Return (row of 'OP', column of month name) in the NANO HFM sheet."""
    full = {1:"january",2:"february",3:"march",4:"april",5:"may",6:"june",
            7:"july",8:"august",9:"september",10:"october",11:"november",12:"december"}
    op_row: Optional[int] = None
    for idx, row in df.iterrows():
        if "op" in row.astype(str).str.lower().values:
            op_row = idx
            break
    month_col: Optional[int] = None
    name = full[month_number]
    for ci, col_name in enumerate(df.columns):
        if name in col_name.lower():
            month_col = ci
            break
    return op_row, month_col
 
 
def find_outside_profit_and_month_value(
    df: pd.DataFrame, month_number: int
) -> tuple[Optional[int], Optional[int]]:
    adj = (month_number - 4) % 12
    for row_idx, row in df.iterrows():
        for ci, val in enumerate(row):
            if str(val).replace(" ", "").lower() == "outsideprofit":
                return row_idx + 1, ci + adj
    return None, None
 
 
def find_second_column_for_outside_profit(
    df: pd.DataFrame, month_number: int
) -> tuple[Optional[int], Optional[int]]:
    adj = (month_number - 4) % 12
    count = 0
    for row_idx, row in df.iterrows():
        for ci, val in enumerate(row):
            if str(val).replace(" ", "").lower() == "outsideprofit":
                count += 1
                if count == 2:
                    return row_idx + 1, ci + adj
    return None, None
 
 
# ══════════════════════════════════════════════════════════════════════════════
#  Section 3 – Excel formula builders
# ══════════════════════════════════════════════════════════════════════════════
 
def _col_letter(col_num: int) -> str:
    """Convert 1-based column number to Excel letter(s).  1→A, 27→AA, etc."""
    letter = ""
    while col_num > 0:
        col_num, rem = divmod(col_num - 1, 26)
        letter = chr(65 + rem) + letter
    return letter
 
 
def generate_excel_formula(
    row_num: int,
    col_num: int,
    sheet_name: Optional[str] = None,
    file_name: Optional[str] = None,
) -> str:
    """
    Build an Excel cell-reference formula string.
 
    Parameters
    ----------
    row_num   : 0-based pandas row index  → Excel row = row_num + 2
    col_num   : 0-based pandas col index  → Excel col = col_num + 1
    sheet_name: worksheet name (optional)
    file_name : full path of the source workbook (optional).
                Backslashes are converted to forward-slashes for Excel
                cross-workbook formula compatibility.
 
    Returns a string starting with '=' ready to be written into a cell.
 
    LINK SAFETY
    -----------
    When file_name is supplied the formula includes the full path so that
    the link is ABSOLUTE.  openpyxl writes this verbatim and Excel resolves
    it correctly when the output file is opened from any location, provided
    the source file hasn't moved.
    """
    excel_row = row_num + 2
    excel_col = col_num + 1
    col_letter = _col_letter(excel_col)
 
    if file_name and sheet_name:
        # Normalise path separators → forward slash (Excel cross-book links)
        safe_path = file_name.replace("\\", "/")
        return f"='[{safe_path}]{sheet_name}'!{col_letter}{excel_row}"
    elif sheet_name:
        return f"='{sheet_name}'!{col_letter}{excel_row}"
    else:
        return f"={col_letter}{excel_row}"
 
 
def generate_excel_formula_249(month_number: int, file_path: str) -> str:
    """
    Build the SUMIF formula for row 249 (Infiniti OSP line).
    Column is derived from month_number relative to April.
    """
    # Column index: AZ = 52, AA = 27 … build offset from April baseline
    base_col = 92   # = 66 = BN in 1-based; original logic kept
    col_idx  = base_col + month_number - 4
    col_let  = _col_letter(col_idx)
    return (
        f"=SUMIF('Master'!$X:$X,\"Infiniti\",'Master'!{col_let}:{col_let})"
        f"/'Master'!{col_let}$1*1000"
    )
 
 
def extract_formula(
    file_path: str,
    sheet_name: str,
    row_number: int,
    column_index: int,
) -> Optional[str]:
    """
    Read a formula (or value) from an existing workbook cell.
 
    Uses data_only=False so we get the formula text, not the cached result.
    keep_links=True prevents openpyxl from breaking existing external refs
    during the read (even though we do not save here).
    """
    try:
        wb = _cached_workbook(file_path)
        ws = wb[sheet_name]
        cell = ws.cell(row=row_number, column=column_index + 1)
        return cell.value  # formula string if the cell has one
    except KeyError:
        raise ValueError(f"Sheet '{sheet_name}' not found in '{file_path}'.")
    except Exception as e:
        raise RuntimeError(f"extract_formula error ({file_path}, {sheet_name}, "
                           f"row={row_number}, col={column_index}): {e}") from e
 
 
def extract_value_formula_latest(file_name: str, month_number: int) -> str:
    """
    Return a formula string referencing cell V8 (0-based: row 6, col 21)
    in the HFM sheet for *month_number*.
    For months after April the formula is a difference vs the previous month's sheet.
    """
    abbr = {1:"Jan",2:"Feb",3:"Mar",4:"Apr",5:"May",6:"Jun",
            7:"Jul",8:"Aug",9:"Sep",10:"Oct",11:"Nov",12:"Dec"}
    sheet_curr = f"HFM {abbr[month_number]}"
    safe_path  = file_name.replace("\\", "/")
    col_letter = _col_letter(22)   # col index 21 (0-based) → 22 (1-based)
    excel_row  = 8                 # row 6 (0-based) → row 8 (1-based, offset +2)
 
    formula_curr = f"='[{safe_path}]{sheet_curr}'!{col_letter}{excel_row}"
    if month_number == 4:
        return formula_curr
 
    prev_num   = month_number - 1 if month_number > 1 else 12
    sheet_prev = f"HFM {abbr[prev_num]}"
    formula_prev = f"='[{safe_path}]{sheet_prev}'!{col_letter}{excel_row}"
    return f"(({formula_curr[1:]} - {formula_prev[1:]})"
 
 
def manage_sheet_visibility(file_name: str, month_number: int) -> str:
    """Unhide current month's HFM sheet and hide the previous month's."""
    abbr = {1:"Jan",2:"Feb",3:"Mar",4:"Apr",5:"May",6:"Jun",
            7:"Jul",8:"Aug",9:"Sep",10:"Oct",11:"Nov",12:"Dec"}
    curr_sheet = f"HFM {abbr[month_number]}"
    prev_num   = month_number - 1 if month_number > 1 else 12
    prev_sheet = f"HFM {abbr[prev_num]}"
 
    wb = openpyxl.load_workbook(file_name, keep_links=True)
    if curr_sheet in wb.sheetnames:
        wb[curr_sheet].sheet_state = "visible"
    else:
        raise ValueError(f"Sheet '{curr_sheet}' not found.")
    if prev_sheet in wb.sheetnames:
        wb[prev_sheet].sheet_state = "hidden"
    wb.save(file_name)
    return f"Unhid '{curr_sheet}', hid '{prev_sheet}'."
 
 
# ══════════════════════════════════════════════════════════════════════════════
#  Section 4 – Main writer: paste_values_KPI_PL
# ══════════════════════════════════════════════════════════════════════════════
 
def paste_values_KPI_PL(
    original_file_path: str,
    sheet_names: list[str],
    cell_updates_list: list[list[tuple]],
    month_name: int,
    output_path: str,
) -> tuple[str, str]:
    """
    Write formulas / values into the copied template workbook and save it.
 
    LINK PRESERVATION STRATEGY
    ---------------------------
    1.  keep_links=True  → openpyxl retains the [xl/externalLinks] part of
        the OOXML package.  Without this flag openpyxl silently removes all
        cross-workbook relationship entries and Excel shows '#REF!' on every
        external formula when the saved file is opened.
 
    2.  We write formula strings (those starting with '=') directly as cell
        values.  openpyxl serialises them as <f>…</f> nodes.  Excel recalc-
        ulates them on first open.  We never write cached values (data_only).
 
    3.  We do NOT use keep_vba=True here because the template is a plain xlsx.
        If your template is xlsm (has macros) change this to keep_vba=True.
 
    Parameters
    ----------
    original_file_path  : path to the copied + Master-injected template
    sheet_names         : ordered list of sheet names matching cell_updates_list
    cell_updates_list   : list of [(row, col, value_or_formula), …] per sheet
    month_name          : 1-based month integer (used for the output filename)
    output_path         : directory where the result file is saved
 
    Returns
    -------
    (legacy_path_unused, full_output_path)
    """
    wb = openpyxl.load_workbook(original_file_path, keep_links=True)

    for sheet_name, updates in zip(sheet_names, cell_updates_list):
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not in workbook '{original_file_path}'.")
        ws = wb[sheet_name]
        for row, col, value in updates:
            if value is None:
                continue
            cell = ws.cell(row=row, column=col)
            # Write formula string directly; write non-string scalars as-is
            cell.value = value

    month_label = calendar.month_name[month_name]
    out_name    = f"FY25 P&L Summary {month_label}_output_FY{CURRENT_FY}.xlsx"
    os.makedirs(output_path, exist_ok=True)
    out_full    = os.path.join(output_path, out_name)

    # Save ONCE to the real output location (no duplicate/legacy copies).
    wb.save(out_full)
    wb.close()

    # Register the cross-workbook links openpyxl wrote as raw strings so Excel
    # opens the file without the "recover unreadable content / repair links"
    # dialog and the links resolve correctly.
    try:
        repaired = repair_external_links(out_full)
        print(f"Saved output → {out_full}  (links repaired: {repaired})")
    except Exception as e:  # never let link-repair break the run
        print(f"[WARN] link repair skipped for {out_full}: {e}")

    # Return the single output path for BOTH legacy slots so existing callers
    # (which unpack two values) keep working without a second physical file.
    return out_full, out_full
 
 
 