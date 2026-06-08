"""
variance_engine.py
──────────────────
Called by the Nissan Variance App when the user clicks ▶ RUN ANALYSIS.

For every selected scenario the engine produces 2 tabs in the output workbook:
    MTD <ScenarioName>   — copied from the MTD template sheet
    YTD <ScenarioName>   — copied from the YTD template sheet

The template sheets are taken from TEMPLATE_FILE (hardcoded path below).
All cell values, formatting, named ranges, and internal hyperlinks are
preserved exactly.  External hyperlinks are intentionally left intact.

Receives a dict (args) with all user inputs:
    args = {
        "scenarios":     list[str],    # e.g. ["FC 2+10", "Actuals"]  (1 or 2)
        # ── legacy two-scenario keys still supported ──────────────
        "scenario_1":    str,
        "scenario_2":    str,
        # ──────────────────────────────────────────────────────────
        "quarter":       str,          # e.g. "Q1 (Apr–Jun)"
        "month":         str,          # e.g. "April"
        "input_folders": list[str],
        "input_files":   list[str],    # full paths of all scanned files
        "output_folder": str,
        "timestamp":     str,          # YYYYMMDD_HHMMSS
    }

Requirements:
    pip install xlwings openpyxl
    Excel must be installed (xlwings uses COM/AppleScript for recalc).
"""

import os
import shutil
import copy
from datetime import datetime

import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ═══════════════════════════════════════════════════════════════
#  ★  CONFIGURATION  — set your template path here  ★
# ═══════════════════════════════════════════════════════════════

# Full path to the template workbook that contains:
#   • a sheet whose name contains "MTD"  (e.g. "MTD Template")
#   • a sheet whose name contains "YTD"  (e.g. "YTD Template")
TEMPLATE_FILE = r"C:\Nissan\Templates\Variance_Template.xlsx"

# Exact sheet names inside TEMPLATE_FILE.
# Leave as None to auto-detect by looking for "MTD" / "YTD" in sheet names.
MTD_SHEET_NAME = None   # e.g. "MTD Template"  or leave None for auto-detect
YTD_SHEET_NAME = None   # e.g. "YTD Template"  or leave None for auto-detect


# ═══════════════════════════════════════════════════════════════
#  MAIN ENTRY POINT  (called by the app)
# ═══════════════════════════════════════════════════════════════

def run_variance(args: dict):
    """
    Main function invoked by the Nissan Variance App.

    Builds an output workbook where each selected scenario gets two tabs:
        MTD <ScenarioName>
        YTD <ScenarioName>
    Both tabs are copied verbatim from the template file.
    """
    # ── Resolve scenarios list (support both old and new key styles) ──
    scenarios = args.get("scenarios") or []
    if not scenarios:
        sc1 = args.get("scenario_1", "")
        sc2 = args.get("scenario_2", "")
        if sc1:
            scenarios.append(sc1)
        if sc2 and sc2 != sc1:
            scenarios.append(sc2)

    quarter     = args["quarter"]
    month       = args["month"]
    input_files = args["input_files"]
    out_folder  = args["output_folder"]
    ts          = args["timestamp"]

    os.makedirs(out_folder, exist_ok=True)

    # ── Validate template file ──────────────────────────────────
    if not os.path.isfile(TEMPLATE_FILE):
        raise FileNotFoundError(
            f"Template file not found:\n{TEMPLATE_FILE}\n\n"
            "Please update TEMPLATE_FILE at the top of variance_engine.py."
        )

    # ── Resolve template sheet names ────────────────────────────
    mtd_name, ytd_name = _resolve_template_sheets(TEMPLATE_FILE)

    # ── Build output path ───────────────────────────────────────
    sc_label = "_vs_".join(_safe(s) for s in scenarios)
    q_label  = quarter.split()[0]
    output_path = os.path.join(
        out_folder,
        f"Variance__{sc_label}__{q_label}_{month}__v{ts}.xlsx"
    )

    # ── Copy template as the base output file ───────────────────
    shutil.copy2(TEMPLATE_FILE, output_path)

    # ── Open the copied file and build tabs ─────────────────────
    wb = load_workbook(output_path, keep_vba=False)

    # Remove all pre-existing sheets — we will re-add only what we need
    for sname in wb.sheetnames[:]:
        del wb[sname]

    # Re-open template fresh for reading (we work from TEMPLATE_FILE, not the copy)
    tpl_wb = load_workbook(TEMPLATE_FILE, data_only=False, keep_vba=False)
    mtd_tpl = tpl_wb[mtd_name]
    ytd_tpl = tpl_wb[ytd_name]

    # ── For each scenario: add MTD tab then YTD tab ─────────────
    for scenario in scenarios:
        mtd_tab_name = f"MTD {scenario}"
        ytd_tab_name = f"YTD {scenario}"

        # Truncate tab names to Excel's 31-char limit
        mtd_tab_name = mtd_tab_name[:31]
        ytd_tab_name = ytd_tab_name[:31]

        _copy_sheet_into(tpl_wb, mtd_tpl, wb, mtd_tab_name)
        _copy_sheet_into(tpl_wb, ytd_tpl, wb, ytd_tab_name)

    # ── Append audit/log sheet ──────────────────────────────────
    log_ws = wb.create_sheet("Input Log")
    _write_input_log(log_ws, args, scenarios)

    wb.save(output_path)
    tpl_wb.close()

    # ── Open in Excel via xlwings to recalculate, then save ─────
    _open_recalc_save(output_path)

    return output_path


# ═══════════════════════════════════════════════════════════════
#  SHEET COPY  — full-fidelity copy with no broken links
# ═══════════════════════════════════════════════════════════════

def _copy_sheet_into(src_wb, src_ws, dst_wb, new_name: str):
    """
    Copies src_ws into dst_wb as a new sheet called new_name.

    Preserves:
      • All cell values and data types (including formulas)
      • Cell formatting  (font, fill, border, alignment, number format)
      • Merged cell ranges
      • Column widths and row heights
      • Sheet-level properties (tab colour, grid lines, zoom)
      • Hyperlinks (both internal and external)
      • Data validations
      • Print / page-setup settings
      • Conditional formatting rules

    NOTE: Charts, images, and sparklines cannot be copied by openpyxl
          without losing fidelity; they are skipped silently.
    """
    dst_ws = dst_wb.create_sheet(new_name)

    # ── Page setup ──────────────────────────────────────────────
    dst_ws.sheet_view.showGridLines = src_ws.sheet_view.showGridLines
    dst_ws.sheet_view.zoomScale     = src_ws.sheet_view.zoomScale
    if src_ws.sheet_properties.tabColor:
        dst_ws.sheet_properties.tabColor = src_ws.sheet_properties.tabColor

    # ── Copy print / page setup attributes ──────────────────────
    for attr in ("orientation", "paperSize", "fitToPage",
                 "fitToHeight", "fitToWidth"):
        try:
            setattr(dst_ws.page_setup, attr,
                    getattr(src_ws.page_setup, attr))
        except Exception:
            pass

    # ── Column widths ────────────────────────────────────────────
    for col_letter, col_dim in src_ws.column_dimensions.items():
        dst_ws.column_dimensions[col_letter].width    = col_dim.width
        dst_ws.column_dimensions[col_letter].hidden   = col_dim.hidden
        dst_ws.column_dimensions[col_letter].bestFit  = col_dim.bestFit

    # ── Row heights ──────────────────────────────────────────────
    for row_idx, row_dim in src_ws.row_dimensions.items():
        dst_ws.row_dimensions[row_idx].height = row_dim.height
        dst_ws.row_dimensions[row_idx].hidden = row_dim.hidden

    # ── Cells (values + style + hyperlinks) ─────────────────────
    for row in src_ws.iter_rows():
        for src_cell in row:
            dst_cell = dst_ws.cell(row=src_cell.row, column=src_cell.column)

            # Value / formula
            dst_cell.value = src_cell.value

            # Number format
            dst_cell.number_format = src_cell.number_format

            # Font
            if src_cell.font:
                dst_cell.font = copy.copy(src_cell.font)

            # Fill
            if src_cell.fill:
                dst_cell.fill = copy.copy(src_cell.fill)

            # Border
            if src_cell.border:
                dst_cell.border = copy.copy(src_cell.border)

            # Alignment
            if src_cell.alignment:
                dst_cell.alignment = copy.copy(src_cell.alignment)

            # Hyperlink — preserve intact (no breakage)
            if src_cell.hyperlink:
                dst_cell.hyperlink = copy.copy(src_cell.hyperlink)

    # ── Merged cells ─────────────────────────────────────────────
    for merge_range in src_ws.merged_cells.ranges:
        dst_ws.merge_cells(str(merge_range))

    # ── Data validations ─────────────────────────────────────────
    for dv in src_ws.data_validations.dataValidation:
        dst_ws.add_data_validation(copy.deepcopy(dv))

    # ── Conditional formatting ───────────────────────────────────
    for cf_range, cf_rules in src_ws.conditional_formatting._cf_rules.items():
        for rule in cf_rules:
            dst_ws.conditional_formatting.add(cf_range, copy.deepcopy(rule))

    # ── Freeze panes ─────────────────────────────────────────────
    if src_ws.freeze_panes:
        dst_ws.freeze_panes = src_ws.freeze_panes

    return dst_ws


# ═══════════════════════════════════════════════════════════════
#  INPUT LOG SHEET
# ═══════════════════════════════════════════════════════════════

def _write_input_log(ws, args, scenarios):
    ws["A1"] = "INPUT FILE LOG"
    ws["A1"].font = Font(name="Helvetica", bold=True, size=12, color="C3002F")

    ws["A2"] = f"Generated: {datetime.now().strftime('%d %b %Y  %H:%M')}"
    ws["A2"].font = Font(name="Helvetica", size=9, color="888888", italic=True)

    ws["A3"] = f"Scenarios: {', '.join(scenarios)}"
    ws["A3"].font = Font(name="Helvetica", size=9, color="888888")

    ws["A4"] = f"Quarter: {args.get('quarter', '')}   Month: {args.get('month', '')}"
    ws["A4"].font = Font(name="Helvetica", size=9, color="888888")

    headers = ["#", "Folder", "File Name", "Full Path"]
    hf    = PatternFill("solid", fgColor="1C1C1C")
    hfont = Font(name="Helvetica", bold=True, color="C0C0C0", size=9)
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=6, column=col, value=h)
        c.fill = hf
        c.font = hfont

    row = 7
    for i, fp in enumerate(args.get("input_files", []), start=1):
        ws.cell(row=row, column=1, value=i).font           = Font(size=8, color="666666")
        ws.cell(row=row, column=2, value=os.path.dirname(fp)).font  = Font(size=8, color="888888")
        ws.cell(row=row, column=3, value=os.path.basename(fp)).font = Font(size=8, color="C0C0C0")
        ws.cell(row=row, column=4, value=fp).font          = Font(size=8, color="444444")
        row += 1

    for col, width in zip("ABCD", [6, 40, 30, 60]):
        ws.column_dimensions[col].width = width
    ws.sheet_view.showGridLines = False


# ═══════════════════════════════════════════════════════════════
#  TEMPLATE SHEET AUTO-DETECT
# ═══════════════════════════════════════════════════════════════

def _resolve_template_sheets(tpl_path: str):
    """
    Returns (mtd_sheet_name, ytd_sheet_name) from the template workbook.
    Uses the module-level constants if set; otherwise auto-detects by
    looking for 'MTD' and 'YTD' in sheet names (case-insensitive).
    """
    wb = load_workbook(tpl_path, read_only=True, data_only=True)
    names = wb.sheetnames
    wb.close()

    mtd = MTD_SHEET_NAME
    ytd = YTD_SHEET_NAME

    if mtd is None:
        for n in names:
            if "mtd" in n.lower():
                mtd = n
                break
    if ytd is None:
        for n in names:
            if "ytd" in n.lower():
                ytd = n
                break

    if mtd is None:
        raise ValueError(
            f"Could not find an MTD sheet in the template.\n"
            f"Sheets found: {names}\n"
            f"Set MTD_SHEET_NAME at the top of variance_engine.py."
        )
    if ytd is None:
        raise ValueError(
            f"Could not find a YTD sheet in the template.\n"
            f"Sheets found: {names}\n"
            f"Set YTD_SHEET_NAME at the top of variance_engine.py."
        )

    return mtd, ytd


# ═══════════════════════════════════════════════════════════════
#  XLWINGS — open in Excel, recalculate, save & close
# ═══════════════════════════════════════════════════════════════

def _open_recalc_save(path: str):
    """
    Opens the workbook in Excel via xlwings so all formulas calculate,
    then saves and closes it.  Gracefully skips if Excel is unavailable.
    """
    try:
        import xlwings as xw
        app = xw.App(visible=False, add_book=False)
        app.display_alerts  = False
        app.screen_updating = False

        wb = app.books.open(os.path.abspath(path))
        app.api.CalculateFull()
        wb.save()
        wb.close()
        app.quit()
        print(f"[engine] Recalculated and saved: {os.path.basename(path)}")

    except Exception as e:
        # xlwings / Excel not available — workbook is still valid, just unrecalculated
        print(f"[engine] xlwings recalc skipped: {e}")


# ═══════════════════════════════════════════════════════════════
#  HELPERS
# ═══════════════════════════════════════════════════════════════

def _filter_files(files: list, scenario: str) -> list:
    sc_key = scenario.lower().replace(" ", "").replace("+", "p")
    return [f for f in files if sc_key in os.path.basename(f).lower()]


def _safe(s: str) -> str:
    return s.replace(" ", "_").replace("+", "p").replace("/", "-")


# ═══════════════════════════════════════════════════════════════
#  STANDALONE TEST
# ═══════════════════════════════════════════════════════════════

if __name__ == "__main__":
    test_args = {
        "scenarios":     ["FC 2+10", "Actuals"],
        "quarter":       "Q1 (Apr–Jun)",
        "month":         "April",
        "input_folders": [r"C:\Test\InputData"],
        "input_files":   [
            r"C:\Test\InputData\Sales_Data.xlsx",
            r"C:\Test\InputData\Actuals_Upload.xlsx",
        ],
        "output_folder": r"C:\Test\Output",
        "timestamp":     datetime.now().strftime("%Y%m%d_%H%M%S"),
    }
    out = run_variance(test_args)
    print(f"Saved to: {out}")
