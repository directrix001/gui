"""
variance_engine.py
──────────────────
Called by the Nissan Variance App when the user clicks ▶ RUN ANALYSIS.

Receives a dict (args) with all user inputs:
    args = {
        "scenario_1":    str,          # e.g. "FC 2+10"
        "scenario_2":    str,          # e.g. "Actuals"
        "quarter":       str,          # e.g. "Q1 (Apr–Jun)"
        "month":         str,          # e.g. "April"
        "input_folders": list[str],    # all scanned folders
        "input_files":   list[str],    # full paths of all scanned files
        "output_folder": str,          # where to save output
        "timestamp":     str,          # YYYYMMDD_HHMMSS
    }

Requirements:
    pip install xlwings openpyxl
    Excel must be installed on the machine (xlwings talks to it via COM/AppleScript).
"""

import os
import shutil
from datetime import datetime

import xlwings as xw
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# ═══════════════════════════════════════════════════════════════
#  MAIN ENTRY POINT  (called by the app)
# ═══════════════════════════════════════════════════════════════

def run_variance(args: dict):
    """
    Main function invoked by the Nissan Variance App.
    Builds a variance summary workbook comparing the two selected scenarios.
    """
    sc1          = args["scenario_1"]
    sc2          = args["scenario_2"]
    quarter      = args["quarter"]
    month        = args["month"]
    input_files  = args["input_files"]
    out_folder   = args["output_folder"]
    ts           = args["timestamp"]

    os.makedirs(out_folder, exist_ok=True)

    # ── Step 1: Find the relevant source files ──────────────────
    sc1_files = _filter_files(input_files, sc1)
    sc2_files = _filter_files(input_files, sc2)

    # ── Step 2: Build the output workbook (openpyxl) ───────────
    output_path = os.path.join(
        out_folder,
        f"Variance__{_safe(sc1)}_vs_{_safe(sc2)}__{quarter.split()[0]}_{month}__v{ts}.xlsx"
    )
    _build_variance_workbook(output_path, sc1, sc2, quarter, month,
                              sc1_files, sc2_files, args)

    # ── Step 3: Open in Excel via xlwings & save (triggers recalc) ─
    _open_recalc_save(output_path)

    return output_path


# ═══════════════════════════════════════════════════════════════
#  BUILD WORKBOOK  (openpyxl — no Excel needed at this stage)
# ═══════════════════════════════════════════════════════════════

def _build_variance_workbook(path, sc1, sc2, quarter, month,
                              sc1_files, sc2_files, args):
    wb = openpyxl.Workbook()

    # ── Sheet 1: Cover ──────────────────────────────────────────
    cover = wb.active
    cover.title = "Cover"
    _write_cover(cover, sc1, sc2, quarter, month, args)

    # ── Sheet 2: Variance Summary ───────────────────────────────
    summary = wb.create_sheet("Variance Summary")
    _write_variance_summary(summary, sc1, sc2, sc1_files, sc2_files)

    # ── Sheet 3: Input File Log ─────────────────────────────────
    log = wb.create_sheet("Input Log")
    _write_input_log(log, args)

    wb.save(path)


def _write_cover(ws, sc1, sc2, quarter, month, args):
    # Nissan red header band
    for row in ws.iter_rows(min_row=1, max_row=4, min_col=1, max_col=8):
        for cell in row:
            cell.fill = PatternFill("solid", fgColor="C3002F")

    ws.merge_cells("A1:H4")
    title_cell = ws["A1"]
    title_cell.value = "NISSAN  |  VARIANCE ANALYSIS REPORT"
    title_cell.font = Font(name="Georgia", size=18, bold=True, color="F0F0F0")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Meta rows
    meta = [
        ("Scenario 1",    sc1),
        ("Scenario 2",    sc2),
        ("Quarter",       quarter),
        ("Month",         month),
        ("Generated",     datetime.now().strftime("%d %b %Y  %H:%M")),
        ("Folders",       str(len(args["input_folders"]))),
        ("Total Files",   str(len(args["input_files"]))),
    ]
    label_font  = Font(name="Helvetica", size=10, bold=True, color="888888")
    value_font  = Font(name="Helvetica", size=11, color="F0F0F0")
    row_fill    = PatternFill("solid", fgColor="1C1C1C")

    for i, (label, value) in enumerate(meta, start=6):
        lc = ws.cell(row=i, column=2, value=label)
        vc = ws.cell(row=i, column=4, value=value)
        lc.font = label_font
        vc.font = value_font
        for col in range(1, 9):
            ws.cell(row=i, column=col).fill = row_fill

    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["D"].width = 34
    ws.sheet_view.showGridLines = False


def _write_variance_summary(ws, sc1, sc2, sc1_files, sc2_files):
    # Header
    headers = ["#", "Metric", sc1, sc2, "Variance (Abs)", "Variance (%)", "Flag"]
    header_fill = PatternFill("solid", fgColor="C3002F")
    header_font = Font(name="Helvetica", bold=True, color="F0F0F0", size=10)
    thin = Side(style="thin", color="2A2A2A")
    border = Border(bottom=thin)

    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    # Placeholder rows — real data would be pulled from the source files
    sample_metrics = [
        ("Revenue",           1_200_000, 1_150_000),
        ("Volume (units)",    45_000,    43_200),
        ("NSP per unit",      26.67,     26.62),
        ("Variable Cost",     780_000,   760_000),
        ("Gross Margin",      420_000,   390_000),
        ("Gross Margin %",    0.35,      0.339),
        ("Fixed Overheads",   120_000,   115_000),
        ("EBIT",              300_000,   275_000),
    ]

    alt_fill  = PatternFill("solid", fgColor="1C1C1C")
    alt_fill2 = PatternFill("solid", fgColor="141414")
    green     = Font(name="Helvetica", color="00C853", size=10)
    red_font  = Font(name="Helvetica", color="C3002F", size=10)
    normal    = Font(name="Helvetica", color="C0C0C0", size=10)

    for i, (metric, v1, v2) in enumerate(sample_metrics, start=2):
        fill = alt_fill if i % 2 == 0 else alt_fill2
        row_data = [i - 1, metric, v1, v2,
                    f"={_col(3)}{i}-{_col(4)}{i}",          # Variance abs (formula)
                    f"=IF({_col(4)}{i}<>0,({_col(3)}{i}-{_col(4)}{i})/{_col(4)}{i},\"N/A\")",
                    f"=IF({_col(6)}{i}>0.05,\"▲ FAV\",IF({_col(6)}{i}<-0.05,\"▼ UNF\",\"—\"))"]

        for col, val in enumerate(row_data, start=1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.fill = fill
            cell.font = normal
            cell.alignment = Alignment(horizontal="center" if col != 2 else "left")

        # Colour the flag column based on formula result at runtime
        # (will resolve when Excel recalcs via xlwings)

    # Source file notes
    ws.cell(row=len(sample_metrics) + 4, column=1,
            value=f"Source ({sc1}): {', '.join(os.path.basename(f) for f in sc1_files) or 'N/A'}"
            ).font = Font(name="Helvetica", size=8, color="666666", italic=True)
    ws.cell(row=len(sample_metrics) + 5, column=1,
            value=f"Source ({sc2}): {', '.join(os.path.basename(f) for f in sc2_files) or 'N/A'}"
            ).font = Font(name="Helvetica", size=8, color="666666", italic=True)

    # Column widths
    for col, width in zip("ABCDEFG", [6, 28, 18, 18, 18, 16, 12]):
        ws.column_dimensions[col].width = width
    ws.sheet_view.showGridLines = False


def _write_input_log(ws, args):
    ws["A1"] = "INPUT FILE LOG"
    ws["A1"].font = Font(name="Helvetica", bold=True, size=12, color="C3002F")

    headers = ["#", "Folder", "File Name", "Full Path"]
    hf = PatternFill("solid", fgColor="1C1C1C")
    hfont = Font(name="Helvetica", bold=True, color="C0C0C0", size=9)
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=3, column=col, value=h)
        c.fill = hf
        c.font = hfont

    row = 4
    for i, fp in enumerate(args["input_files"], start=1):
        ws.cell(row=row, column=1, value=i).font = Font(size=8, color="666666")
        ws.cell(row=row, column=2, value=os.path.dirname(fp)).font = Font(size=8, color="888888")
        ws.cell(row=row, column=3, value=os.path.basename(fp)).font = Font(size=8, color="C0C0C0")
        ws.cell(row=row, column=4, value=fp).font = Font(size=8, color="444444")
        row += 1

    for col, width in zip("ABCD", [6, 40, 30, 60]):
        ws.column_dimensions[col].width = width
    ws.sheet_view.showGridLines = False


# ═══════════════════════════════════════════════════════════════
#  XLWINGS — open in Excel, recalculate, save & close
# ═══════════════════════════════════════════════════════════════

def _open_recalc_save(path: str):
    """
    Opens the workbook in Excel via xlwings so all formulas calculate,
    then saves and closes it. Excel must be installed.
    """
    try:
        app = xw.App(visible=False, add_book=False)
        app.display_alerts = False
        app.screen_updating = False

        wb = app.books.open(os.path.abspath(path))

        # Force full recalculation
        app.api.CalculateFull()

        # Example xlwings read after recalc — log gross margin to console
        try:
            ws = wb.sheets["Variance Summary"]
            gm_val  = ws.range("C6").value   # Gross Margin sc1
            gm_pct  = ws.range("F6").value   # Gross Margin %
            print(f"[engine] Gross Margin {gm_val}  |  {gm_pct:.1%}" if gm_pct else "")
        except Exception:
            pass

        wb.save()
        wb.close()
        app.quit()

    except Exception as e:
        # xlwings / Excel not available — workbook is still valid, just unrecalculated
        print(f"[engine] xlwings recalc skipped: {e}")


# ═══════════════════════════════════════════════════════════════
#  HELPERS
# ═══════════════════════════════════════════════════════════════

def _filter_files(files: list, scenario: str) -> list:
    """Return files whose names loosely match the scenario label."""
    sc_key = scenario.lower().replace(" ", "").replace("+", "p")
    return [f for f in files if sc_key in os.path.basename(f).lower()]


def _safe(s: str) -> str:
    return s.replace(" ", "_").replace("+", "p").replace("/", "-")


def _col(n: int) -> str:
    """Convert column number to Excel letter (1→A, 3→C …)."""
    result = ""
    while n:
        n, r = divmod(n - 1, 26)
        result = chr(65 + r) + result
    return result


# ═══════════════════════════════════════════════════════════════
#  STANDALONE TEST
# ═══════════════════════════════════════════════════════════════

if __name__ == "__main__":
    test_args = {
        "scenario_1":    "FC 2+10",
        "scenario_2":    "Actuals",
        "quarter":       "Q1 (Apr–Jun)",
        "month":         "April",
        "input_folders": [r"C:\Test\InputData"],
        "input_files":   [
            r"C:\Test\InputData\Sales_Data.xlsx",
            r"C:\Test\InputData\Actuals_Upload.xlsx",
        ],
        "output_folder": r"C:\Test\Output",
        "timestamp":     datetime.now().strftime("%Y%m%d_%H%M%S"),
    }
    out = run_variance(test_args)
    print(f"Saved to: {out}")
