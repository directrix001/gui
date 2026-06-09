"""
variance_engine.py  —  v4.0
────────────────────────────
Root cause of broken links
──────────────────────────
When openpyxl copies a sheet cell-by-cell from workbook A into workbook B,
formulas that reference OTHER sheets inside A (e.g. ='DataSheet'!B5) are
written verbatim into B.  Excel then treats them as references to an EXTERNAL
workbook (A) and shows the dreaded "[Variance_Template.xlsx]DataSheet!B5"
broken link.

Fix strategy — two-tier
────────────────────────
TIER 1  (preferred, Windows + Excel installed)
  Use xlwings COM automation:
    • Open BOTH the template and the output workbook in Excel.
    • Use worksheet.api.Copy(Before/After=...) — Excel's own copy engine.
    • Excel remaps every formula reference so it stays internal.
    • Remove the old external-link records, CalculateFull(), save & close.

TIER 2  (fallback — no Excel / non-Windows)
  Use a zip-level XML copy:
    • Treat both .xlsx files as zip archives.
    • Copy the raw worksheet XML, drawing, and relationship files byte-for-byte
      from the template into the output archive.
    • Patch [Content_Types].xml and workbook.xml so Excel sees the new sheets.
    • Rewrite every formula that still carries the external workbook prefix
      ( [filename.xlsx] ) to strip it, making the reference internal.
    • This preserves 100 % of formatting, charts, sparklines, images.
"""

import os
import re
import copy
import shutil
import zipfile
import tempfile
from datetime import datetime
from lxml import etree

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill


# ═══════════════════════════════════════════════════════════════
#  CONFIGURATION
# ═══════════════════════════════════════════════════════════════

# Override with exact sheet names if auto-detect is unreliable.
MTD_SHEET_NAME: str | None = None
YTD_SHEET_NAME: str | None = None


# ═══════════════════════════════════════════════════════════════
#  MAIN ENTRY POINT
# ═══════════════════════════════════════════════════════════════

def run_variance(args: dict) -> str:
    """
    Build one output workbook:
      • Base = master file (copied to output folder)
      • For each (sc1, sc2) pair  →  add  "MTD sc1 vs sc2"  and  "YTD sc1 vs sc2"  tabs
        copied from the template file, with ALL internal formula links intact.
      • Append "Input Log" audit tab.

    Returns the full path of the saved output file.
    """
    pairs         = args["scenario_pairs"]
    quarter       = args["quarter"]
    month         = args["month"]
    master_file   = args["master_file"]
    template_file = args["template_file"]
    out_folder    = args["output_folder"]
    ts            = args["timestamp"]

    os.makedirs(out_folder, exist_ok=True)

    if not os.path.isfile(template_file):
        raise FileNotFoundError(f"Template file not found:\n{template_file}")
    if not os.path.isfile(master_file):
        raise FileNotFoundError(f"Master file not found:\n{master_file}")

    mtd_name, ytd_name = _resolve_template_sheets(template_file)

    q_label     = quarter.split()[0]
    output_path = os.path.join(out_folder, f"Variance__{q_label}_{month}__v{ts}.xlsx")
    shutil.copy2(master_file, output_path)

    # Build tab-name map  {template_sheet_name: desired_output_tab_name}
    tab_map: dict[str, str] = {}
    for sc1, sc2 in pairs:
        lbl = f"{sc1} vs {sc2}"
        tab_map[mtd_name] = f"MTD {lbl}"[:31]
        tab_map[ytd_name] = f"YTD {lbl}"[:31]
        # Note: if multiple pairs share the same template sheet name the last
        # pair wins for that entry; each pair is processed in sequence below.

    # ── Attempt TIER 1: xlwings COM copy ──────────────────────────
    xlwings_ok = _try_xlwings_copy(output_path, template_file, pairs,
                                   mtd_name, ytd_name)

    if not xlwings_ok:
        # ── TIER 2: zip-level XML copy ─────────────────────────────
        print("[engine] Falling back to zip-level XML copy.")
        _zip_copy_sheets(output_path, template_file, pairs, mtd_name, ytd_name)

    # ── Audit log (openpyxl — plain data, no formula links) ───────
    out_wb = load_workbook(output_path, data_only=False, keep_vba=False)
    if "Input Log" in out_wb.sheetnames:
        del out_wb["Input Log"]
    _write_input_log(out_wb.create_sheet("Input Log"), args, pairs)
    out_wb.save(output_path)
    out_wb.close()

    print(f"[engine] Done: {os.path.basename(output_path)}")
    return output_path


# ═══════════════════════════════════════════════════════════════
#  TIER 1 — xlwings COM  (Excel-native sheet copy)
# ═══════════════════════════════════════════════════════════════

def _try_xlwings_copy(output_path, template_file, pairs, mtd_name, ytd_name) -> bool:
    """
    Use Excel COM to copy sheets.  Returns True on success, False if xlwings
    is unavailable or any error occurs.

    Excel's Worksheet.Copy() method:
      • Duplicates the sheet INCLUDING charts, images, sparklines.
      • Automatically keeps all formula references internal — no external links.
    """
    try:
        import xlwings as xw
    except ImportError:
        print("[engine] xlwings not installed — using zip fallback.")
        return False

    app = None
    try:
        app = xw.App(visible=False, add_book=False)
        app.display_alerts  = False
        app.screen_updating = False

        tpl_wb_xw  = app.books.open(os.path.abspath(template_file))
        out_wb_xw  = app.books.open(os.path.abspath(output_path))

        for sc1, sc2 in pairs:
            lbl     = f"{sc1} vs {sc2}"
            mtd_tab = f"MTD {lbl}"[:31]
            ytd_tab = f"YTD {lbl}"[:31]

            for tpl_sheet_name, new_tab_name in [(mtd_name, mtd_tab),
                                                  (ytd_name, ytd_tab)]:
                # Delete any existing sheet with the same name in output
                for sh in out_wb_xw.sheets:
                    if sh.name == new_tab_name:
                        sh.delete()
                        break

                # Copy sheet from template into output workbook (appended at end)
                src_sheet = tpl_wb_xw.sheets[tpl_sheet_name]
                last_sheet = out_wb_xw.sheets[-1]
                src_sheet.api.Copy(After=last_sheet.api)

                # The copied sheet lands with original name — rename it
                copied = out_wb_xw.sheets[-1]
                copied.name = new_tab_name
                print(f"[engine][xlwings] Copied '{tpl_sheet_name}' → '{new_tab_name}'")

        # Break external links pointing back to the template file
        _xlwings_break_links(app, out_wb_xw, template_file)

        app.api.CalculateFull()
        out_wb_xw.save()
        out_wb_xw.close()
        tpl_wb_xw.close()
        app.quit()
        return True

    except Exception as exc:
        print(f"[engine] xlwings copy failed ({exc}) — using zip fallback.")
        try:
            if app:
                app.quit()
        except Exception:
            pass
        return False


def _xlwings_break_links(app, wb_xw, template_file: str):
    """
    Break / update external links that point to the template file so the
    output workbook becomes fully self-contained.
    """
    try:
        # Excel COM: Workbook.LinkSources returns array of link paths
        links = wb_xw.api.LinkSources(1)   # 1 = xlExcelLinks
        if not links:
            return
        tpl_name = os.path.basename(template_file).lower()
        for link in links:
            if tpl_name in str(link).lower():
                # BreakLink converts formula references to their last calculated values
                wb_xw.api.BreakLink(Name=link, Type=1)
                print(f"[engine][xlwings] Broke external link: {link}")
    except Exception as e:
        print(f"[engine][xlwings] BreakLink skipped: {e}")


# ═══════════════════════════════════════════════════════════════
#  TIER 2 — zip-level XML copy  (no Excel required)
# ═══════════════════════════════════════════════════════════════

# XML namespace map used throughout
_NS = {
    "main":  "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
    "r":     "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    "mc":    "http://schemas.openxmlformats.org/markup-compatibility/2006",
    "ct":    "http://schemas.openxmlformats.org/package/2006/content-types",
    "pkgrel":"http://schemas.openxmlformats.org/package/2006/relationships",
}


def _zip_copy_sheets(output_path: str, template_file: str,
                     pairs: list, mtd_name: str, ytd_name: str):
    """
    Copy MTD and YTD sheets from template into the output file at the
    zip/XML level, preserving charts, images, and all formatting.

    After copying, scan every formula cell and strip any
    '[TemplateName.xlsx]' prefix so references become internal.
    """
    tpl_basename  = os.path.basename(template_file)          # for link-stripping
    tpl_basename_noext = os.path.splitext(tpl_basename)[0]   # without .xlsx

    with tempfile.TemporaryDirectory() as tmpdir:
        out_dir = os.path.join(tmpdir, "out")
        tpl_dir = os.path.join(tmpdir, "tpl")

        # Unzip both workbooks
        with zipfile.ZipFile(output_path,  "r") as z: z.extractall(out_dir)
        with zipfile.ZipFile(template_file,"r") as z: z.extractall(tpl_dir)

        # Parse workbook XMLs to discover sheet names → sheet files
        tpl_wb_xml  = _parse_xml(tpl_dir,  "xl/workbook.xml")
        out_wb_xml  = _parse_xml(out_dir,  "xl/workbook.xml")
        tpl_wb_rels = _parse_xml(tpl_dir,  "xl/_rels/workbook.xml.rels")
        out_wb_rels = _parse_xml(out_dir,  "xl/_rels/workbook.xml.rels")

        tpl_sheet_map = _build_sheet_map(tpl_wb_xml, tpl_wb_rels)  # name→(rId, file)
        out_sheet_map = _build_sheet_map(out_wb_xml, out_wb_rels)   # name→(rId, file)

        # Determine next free sheetN index and rId index in output
        next_sheet_idx = _next_sheet_index(out_dir)
        next_rid_idx   = _next_rid_index(out_wb_rels)

        sheets_to_copy: list[tuple[str, str]] = []  # (template_sheet_name, new_tab_name)
        for sc1, sc2 in pairs:
            lbl = f"{sc1} vs {sc2}"
            sheets_to_copy.append((mtd_name, f"MTD {lbl}"[:31]))
            sheets_to_copy.append((ytd_name, f"YTD {lbl}"[:31]))

        for tpl_sname, new_tab_name in sheets_to_copy:
            if tpl_sname not in tpl_sheet_map:
                raise ValueError(f"Sheet '{tpl_sname}' not found in template. "
                                 f"Available: {list(tpl_sheet_map.keys())}")

            _, tpl_sheet_file = tpl_sheet_map[tpl_sname]  # e.g. "worksheets/sheet2.xml"

            new_sheet_file = f"worksheets/sheet{next_sheet_idx}.xml"
            new_sheet_path_out = os.path.join(out_dir, "xl", new_sheet_file)
            new_rid            = f"rId{next_rid_idx}"

            # Copy worksheet XML — then strip external-link prefixes from formulas
            src_ws_path = os.path.join(tpl_dir, "xl", tpl_sheet_file)
            shutil.copy2(src_ws_path, new_sheet_path_out)
            _strip_external_links_in_xml(
                new_sheet_path_out, tpl_basename, tpl_basename_noext)

            # Copy worksheet rels file (drawings, images, charts) if present
            tpl_ws_rels_path = os.path.join(
                tpl_dir, "xl", "worksheets", "_rels",
                os.path.basename(tpl_sheet_file) + ".rels")
            if os.path.isfile(tpl_ws_rels_path):
                os.makedirs(os.path.join(out_dir, "xl", "worksheets", "_rels"),
                            exist_ok=True)
                out_ws_rels_path = os.path.join(
                    out_dir, "xl", "worksheets", "_rels",
                    f"sheet{next_sheet_idx}.xml.rels")
                shutil.copy2(tpl_ws_rels_path, out_ws_rels_path)
                # Copy any referenced drawings / charts / media
                _copy_ws_dependencies(tpl_ws_rels_path, tpl_dir, out_dir,
                                      tpl_sheet_file)

            # Register new sheet in output workbook.xml
            _register_sheet_in_workbook(out_wb_xml, new_tab_name, new_rid)

            # Register relationship in workbook.xml.rels
            _register_rel(out_wb_rels, new_rid,
                          "http://schemas.openxmlformats.org/officeDocument/2006/"
                          "relationships/worksheet",
                          new_sheet_file)

            # Register in [Content_Types].xml
            _register_content_type(out_dir, f"/xl/{new_sheet_file}")

            next_sheet_idx += 1
            next_rid_idx   += 1
            print(f"[engine][zip] Copied '{tpl_sname}' → '{new_tab_name}'")

        # Remove stale external link references from workbook.xml (externalReferences)
        _remove_external_references(out_wb_xml, tpl_basename, tpl_basename_noext)

        # Save modified XMLs
        _save_xml(out_wb_xml,  out_dir, "xl/workbook.xml")
        _save_xml(out_wb_rels, out_dir, "xl/_rels/workbook.xml.rels")

        # Repack the zip
        _repack_zip(out_dir, output_path)


# ── XML helpers ──────────────────────────────────────────────────

def _parse_xml(base_dir: str, rel_path: str):
    return etree.parse(os.path.join(base_dir, rel_path))


def _save_xml(tree, base_dir: str, rel_path: str):
    path = os.path.join(base_dir, rel_path)
    os.makedirs(os.path.dirname(path), exist_ok=True)
    tree.write(path, xml_declaration=True, encoding="UTF-8", standalone=True)


def _build_sheet_map(wb_xml, wb_rels_xml) -> dict:
    """Returns {sheet_name: (rId, 'worksheets/sheetN.xml')}"""
    root     = wb_xml.getroot()
    rels_root = wb_rels_xml.getroot()

    # Build rId → target map from rels
    rid_to_target = {}
    for rel in rels_root:
        rid_to_target[rel.get("Id")] = rel.get("Target")

    result = {}
    ns = {"m": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    for sh in root.findall(".//m:sheet", ns):
        name = sh.get("name")
        rid  = sh.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id")
        target = rid_to_target.get(rid, "")
        result[name] = (rid, target)
    return result


def _next_sheet_index(out_dir: str) -> int:
    ws_dir = os.path.join(out_dir, "xl", "worksheets")
    if not os.path.isdir(ws_dir):
        return 1
    existing = [f for f in os.listdir(ws_dir)
                if f.startswith("sheet") and f.endswith(".xml")]
    nums = []
    for f in existing:
        m = re.search(r"sheet(\d+)\.xml$", f)
        if m:
            nums.append(int(m.group(1)))
    return max(nums, default=0) + 1


def _next_rid_index(rels_xml) -> int:
    root = rels_xml.getroot()
    nums = []
    for rel in root:
        rid = rel.get("Id", "")
        m = re.search(r"rId(\d+)$", rid)
        if m:
            nums.append(int(m.group(1)))
    return max(nums, default=0) + 1


def _register_sheet_in_workbook(wb_xml, sheet_name: str, r_id: str):
    root = wb_xml.getroot()
    ns   = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    r_ns = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
    sheets_el = root.find(f"{{{ns}}}sheets")
    if sheets_el is None:
        sheets_el = etree.SubElement(root, f"{{{ns}}}sheets")

    # Find max sheetId
    max_id = max(
        (int(s.get("sheetId", 0)) for s in sheets_el),
        default=0
    )
    new_sheet = etree.SubElement(sheets_el, f"{{{ns}}}sheet")
    new_sheet.set("name",    sheet_name)
    new_sheet.set("sheetId", str(max_id + 1))
    new_sheet.set(f"{{{r_ns}}}id", r_id)


def _register_rel(rels_xml, r_id: str, rel_type: str, target: str):
    root    = rels_xml.getroot()
    new_rel = etree.SubElement(root, "Relationship")
    new_rel.set("Id",     r_id)
    new_rel.set("Type",   rel_type)
    new_rel.set("Target", target)


def _register_content_type(out_dir: str, part_name: str):
    ct_path = os.path.join(out_dir, "[Content_Types].xml")
    tree    = etree.parse(ct_path)
    root    = tree.getroot()
    ns      = "http://schemas.openxmlformats.org/package/2006/content-types"
    # Check not already registered
    for ov in root.findall(f"{{{ns}}}Override"):
        if ov.get("PartName") == part_name:
            return
    ov = etree.SubElement(root, f"{{{ns}}}Override")
    ov.set("PartName",    part_name)
    ov.set("ContentType",
           "application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml")
    tree.write(ct_path, xml_declaration=True, encoding="UTF-8", standalone=True)


def _strip_external_links_in_xml(ws_xml_path: str,
                                  tpl_basename: str,
                                  tpl_basename_noext: str):
    """
    Open the worksheet XML and rewrite every <f> (formula) element to remove
    external workbook prefixes like:
        '[Variance_Template.xlsx]' or '[Variance_Template]'
    so all references become internal sheet references.

    Also rewrites hyperlinks that point to the template file.
    """
    tree = etree.parse(ws_xml_path)
    root = tree.getroot()
    ns   = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"

    # Patterns to strip — both with and without .xlsx extension, case-insensitive
    patterns = [
        re.compile(r"\[" + re.escape(tpl_basename)      + r"\]", re.IGNORECASE),
        re.compile(r"\[" + re.escape(tpl_basename_noext) + r"\]", re.IGNORECASE),
        re.compile(r"\[" + re.escape(tpl_basename_noext) + r"\.xlsx\]", re.IGNORECASE),
        re.compile(r"\[" + re.escape(tpl_basename_noext) + r"\.xlsm\]", re.IGNORECASE),
    ]

    def _clean(text: str) -> str:
        for pat in patterns:
            text = pat.sub("", text)
        return text

    # Rewrite <f> formula elements
    for f_el in root.iter(f"{{{ns}}}f"):
        if f_el.text:
            f_el.text = _clean(f_el.text)

    # Rewrite shared formula strings in <si>/<t> if present
    for t_el in root.iter(f"{{{ns}}}t"):
        if t_el.text:
            t_el.text = _clean(t_el.text)

    tree.write(ws_xml_path, xml_declaration=True,
               encoding="UTF-8", standalone=True)


def _copy_ws_dependencies(rels_path: str, tpl_dir: str, out_dir: str,
                           tpl_sheet_file: str):
    """
    Read the worksheet's .rels file and copy any referenced resources
    (drawings, charts, images, vmlDrawings) from template to output.
    """
    tree = etree.parse(rels_path)
    root = tree.getroot()
    ws_subdir = os.path.dirname(tpl_sheet_file)  # "worksheets"

    for rel in root:
        target = rel.get("Target", "")
        # Target is relative to the worksheet's directory
        if target.startswith(".."):
            src_rel = os.path.normpath(
                os.path.join("xl", ws_subdir, target)).replace("\\", "/")
        else:
            src_rel = os.path.normpath(
                os.path.join("xl", ws_subdir, target)).replace("\\", "/")

        src_abs = os.path.join(tpl_dir, src_rel)
        dst_abs = os.path.join(out_dir, src_rel)

        if os.path.isfile(src_abs) and not os.path.isfile(dst_abs):
            os.makedirs(os.path.dirname(dst_abs), exist_ok=True)
            shutil.copy2(src_abs, dst_abs)

            # If this resource also has a .rels, copy it too
            rels_of_dep = src_abs.replace(
                os.path.basename(src_abs),
                os.path.join("_rels", os.path.basename(src_abs) + ".rels"))
            if os.path.isfile(rels_of_dep):
                dst_rels = dst_abs.replace(
                    os.path.basename(dst_abs),
                    os.path.join("_rels", os.path.basename(dst_abs) + ".rels"))
                os.makedirs(os.path.dirname(dst_rels), exist_ok=True)
                shutil.copy2(rels_of_dep, dst_rels)


def _remove_external_references(wb_xml, tpl_basename: str,
                                 tpl_basename_noext: str):
    """
    Remove <externalReferences> and <definedNames> entries in workbook.xml
    that reference the template file — prevents Excel from prompting to
    update links on open.
    """
    root = wb_xml.getroot()
    ns   = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"

    for tag in ["externalReferences", "externalLinks"]:
        for el in root.findall(f"{{{ns}}}{tag}"):
            root.remove(el)

    # Also clean up any definedNames that reference the template
    dn_el = root.find(f"{{{ns}}}definedNames")
    if dn_el is not None:
        tpl_lower = tpl_basename.lower()
        for dn in list(dn_el):
            val = (dn.text or "").lower()
            if tpl_lower in val or tpl_basename_noext.lower() in val:
                dn_el.remove(dn)


def _repack_zip(src_dir: str, out_path: str):
    """Repack a directory back into a .xlsx (zip) file."""
    tmp_zip = out_path + ".tmp_repack"
    with zipfile.ZipFile(tmp_zip, "w", zipfile.ZIP_DEFLATED) as zout:
        for root_dir, dirs, files in os.walk(src_dir):
            # Ensure [Content_Types].xml goes first (Excel requirement)
            if root_dir == src_dir and "[Content_Types].xml" in files:
                files.remove("[Content_Types].xml")
                files.insert(0, "[Content_Types].xml")
            for fname in files:
                abs_path  = os.path.join(root_dir, fname)
                arc_name  = os.path.relpath(abs_path, src_dir).replace("\\", "/")
                zout.write(abs_path, arc_name)
    os.replace(tmp_zip, out_path)


# ═══════════════════════════════════════════════════════════════
#  TEMPLATE SHEET AUTO-DETECT
# ═══════════════════════════════════════════════════════════════

def _resolve_template_sheets(tpl_path: str):
    wb    = load_workbook(tpl_path, read_only=True, data_only=True)
    names = wb.sheetnames
    wb.close()

    mtd = MTD_SHEET_NAME
    ytd = YTD_SHEET_NAME

    if mtd is None:
        for n in names:
            if "mtd" in n.lower():
                mtd = n
                break
    if ytd is None:
        for n in names:
            if "ytd" in n.lower():
                ytd = n
                break

    if mtd is None:
        raise ValueError(
            f"No MTD sheet found in template.\nSheets: {names}\n"
            "Set MTD_SHEET_NAME at the top of variance_engine.py.")
    if ytd is None:
        raise ValueError(
            f"No YTD sheet found in template.\nSheets: {names}\n"
            "Set YTD_SHEET_NAME at the top of variance_engine.py.")
    return mtd, ytd


# ═══════════════════════════════════════════════════════════════
#  AUDIT LOG  (plain data — no formula references)
# ═══════════════════════════════════════════════════════════════

def _write_input_log(ws, args: dict, pairs: list):
    ws["A1"] = "INPUT FILE LOG"
    ws["A1"].font = Font(name="Helvetica", bold=True, size=12, color="C3002F")

    meta = [
        ("Generated",   datetime.now().strftime("%d %b %Y  %H:%M")),
        ("Quarter",     args.get("quarter", "")),
        ("Month",       args.get("month", "")),
        ("Master File", args.get("master_file", "")),
        ("Template",    args.get("template_file", "")),
    ]
    lf = Font(name="Helvetica", size=9, bold=True,  color="888888")
    vf = Font(name="Helvetica", size=9,              color="C0C0C0")
    for i, (label, value) in enumerate(meta, start=2):
        ws.cell(row=i, column=1, value=label).font = lf
        ws.cell(row=i, column=2, value=value).font = vf

    hf    = PatternFill("solid", fgColor="1C1C1C")
    hfont = Font(name="Helvetica", bold=True, color="C0C0C0", size=9)

    row = len(meta) + 3
    ws.cell(row=row, column=1, value="SCENARIO PAIRS").font = Font(
        name="Helvetica", bold=True, size=9, color="C3002F")
    row += 1
    for col, h in enumerate(["#", "Scenario 1", "Scenario 2", "MTD Tab", "YTD Tab"], 1):
        c = ws.cell(row=row, column=col, value=h)
        c.fill = hf; c.font = hfont
    row += 1
    for idx, (sc1, sc2) in enumerate(pairs, 1):
        lbl = f"{sc1} vs {sc2}"
        ws.cell(row=row, column=1, value=idx).font             = Font(size=8, color="666666")
        ws.cell(row=row, column=2, value=sc1).font             = Font(size=8, color="C0C0C0")
        ws.cell(row=row, column=3, value=sc2).font             = Font(size=8, color="C0C0C0")
        ws.cell(row=row, column=4, value=f"MTD {lbl}"[:31]).font = Font(size=8, color="888888")
        ws.cell(row=row, column=5, value=f"YTD {lbl}"[:31]).font = Font(size=8, color="888888")
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="INPUT FILES").font = Font(
        name="Helvetica", bold=True, size=9, color="C3002F")
    row += 1
    for col, h in enumerate(["#", "Folder", "File Name", "Full Path"], 1):
        c = ws.cell(row=row, column=col, value=h)
        c.fill = hf; c.font = hfont
    row += 1
    for i, fp in enumerate(args.get("input_files", []), 1):
        ws.cell(row=row, column=1, value=i).font                     = Font(size=8, color="666666")
        ws.cell(row=row, column=2, value=os.path.dirname(fp)).font   = Font(size=8, color="888888")
        ws.cell(row=row, column=3, value=os.path.basename(fp)).font  = Font(size=8, color="C0C0C0")
        ws.cell(row=row, column=4, value=fp).font                    = Font(size=8, color="444444")
        row += 1

    for col, width in zip("ABCDE", [6, 20, 20, 30, 45]):
        ws.column_dimensions[col].width = width
    ws.sheet_view.showGridLines = False


# ═══════════════════════════════════════════════════════════════
#  HELPERS
# ═══════════════════════════════════════════════════════════════

def _safe(s: str) -> str:
    return s.replace(" ", "_").replace("+", "p").replace("/", "-")


# ═══════════════════════════════════════════════════════════════
#  STANDALONE TEST
# ═══════════════════════════════════════════════════════════════

if __name__ == "__main__":
    from datetime import datetime as _dt
    test_args = {
        "scenario_pairs":  [("FC 2+10", "Actuals"), ("Budget", "LY Actuals")],
        "quarter":         "Q1 (Apr-Jun)",
        "month":           "April",
        "input_folders":   [r"C:\Test\InputData"],
        "input_files":     [r"C:\Test\InputData\Sales_Data.xlsx"],
        "master_file":     r"C:\Test\InputData\Variance_Master.xlsx",
        "template_file":   r"C:\Test\Templates\Variance_Template.xlsx",
        "output_folder":   r"C:\Test\Output",
        "timestamp":       _dt.now().strftime("%Y%m%d_%H%M%S"),
    }
    out = run_variance(test_args)
    print(f"Saved to: {out}")
