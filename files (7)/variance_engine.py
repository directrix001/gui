"""
variance_engine.py
──────────────────
Called by the Nissan Variance App when the user clicks RUN ANALYSIS.

Logic:
  1. Take the 2 selected scenarios (e.g. "FC 2+10" vs "Actuals").
  2. Open the template file (TEMPLATE_FILE — hardcoded path below).
     Auto-detect the MTD and YTD sheets inside it.
  3. Copy the master input file to the output folder as the base workbook.
  4. Into that workbook copy the MTD template sheet  -> "MTD FC 2+10 vs Actuals"
                              and the YTD template sheet  -> "YTD FC 2+10 vs Actuals"
     All formatting, merges, hyperlinks, conditional formatting are preserved intact.
  5. Open in Excel via xlwings to force formula recalculation, then save & close.

Receives a dict (args):
    args = {
        "scenario_1":    str,          # e.g. "FC 2+10"
        "scenario_2":    str,          # e.g. "Actuals"
        "quarter":       str,          # e.g. "Q1 (Apr-Jun)"
        "month":         str,          # e.g. "April"
        "input_folders": list[str],
        "input_files":   list[str],
        "master_file":   str,          # full path selected by user in UI
        "output_folder": str,
        "timestamp":     str,          # YYYYMMDD_HHMMSS
    }

Requirements:
    pip install xlwings openpyxl
    Excel must be installed (xlwings uses COM/AppleScript for recalc).
"""

import os
import copy
import shutil
from datetime import datetime

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# ═══════════════════════════════════════════════════════════════
#  CONFIGURATION  — set your template path here
# ═══════════════════════════════════════════════════════════════

# Full path to the template workbook containing the MTD and YTD sheets.
# This file is read-only; it is never modified.
TEMPLATE_FILE = r"C:\Nissan\Templates\Variance_Template.xlsx"

# Exact sheet names inside TEMPLATE_FILE.
# Leave as None to auto-detect by looking for "MTD" / "YTD" in sheet names.
MTD_SHEET_NAME = None   # e.g. "MTD Template"
YTD_SHEET_NAME = None   # e.g. "YTD Template"


# ═══════════════════════════════════════════════════════════════
#  MAIN ENTRY POINT
# ═══════════════════════════════════════════════════════════════

def run_variance(args: dict) -> str:
    """
    Builds the output workbook for one comparison pair (sc1 vs sc2).

    Output workbook structure:
        MTD <sc1> vs <sc2>    <- copied from MTD template sheet
        YTD <sc1> vs <sc2>    <- copied from YTD template sheet
        Input Log             <- audit trail

    Returns the full path of the saved output file.
    """
    sc1         = args["scenario_1"]
    sc2         = args["scenario_2"]
    quarter     = args["quarter"]
    month       = args["month"]
    master_file = args["master_file"]
    out_folder  = args["output_folder"]
    ts          = args["timestamp"]

    os.makedirs(out_folder, exist_ok=True)

    # ── Validate files ──────────────────────────────────────────
    if not os.path.isfile(TEMPLATE_FILE):
        raise FileNotFoundError(
            f"Template file not found:\n{TEMPLATE_FILE}\n\n"
            "Update TEMPLATE_FILE at the top of variance_engine.py."
        )
    if not os.path.isfile(master_file):
        raise FileNotFoundError(
            f"Master input file not found:\n{master_file}"
        )

    # ── Resolve MTD / YTD sheet names from template ─────────────
    mtd_name, ytd_name = _resolve_template_sheets(TEMPLATE_FILE)

    # ── Tab names for the output workbook ───────────────────────
    pair_label  = f"{sc1} vs {sc2}"
    mtd_tab     = f"MTD {pair_label}"[:31]   # Excel 31-char sheet name limit
    ytd_tab     = f"YTD {pair_label}"[:31]

    # ── Build output file path ───────────────────────────────────
    sc_safe = _safe(sc1) + "_vs_" + _safe(sc2)
    q_label = quarter.split()[0]
    output_path = os.path.join(
        out_folder,
        f"Variance__{sc_safe}__{q_label}_{month}__v{ts}.xlsx"
    )

    # ── Copy master file as the base output workbook ─────────────
    # The master file carries the data; we ADD the template tabs to it.
    shutil.copy2(master_file, output_path)

    # ── Open output workbook and inject template tabs ─────────────
    out_wb  = load_workbook(output_path, data_only=False, keep_vba=False)
    tpl_wb  = load_workbook(TEMPLATE_FILE, data_only=False, keep_vba=False)

    mtd_src = tpl_wb[mtd_name]
    ytd_src = tpl_wb[ytd_name]

    # Remove any pre-existing sheets with the same names (idempotent re-runs)
    for name in [mtd_tab, ytd_tab, "Input Log"]:
        if name in out_wb.sheetnames:
            del out_wb[name]

    # Copy template sheets into output workbook
    _copy_sheet_into(tpl_wb, mtd_src, out_wb, mtd_tab)
    _copy_sheet_into(tpl_wb, ytd_src, out_wb, ytd_tab)

    # Append audit log
    _write_input_log(out_wb.create_sheet("Input Log"), args, sc1, sc2)

    out_wb.save(output_path)
    tpl_wb.close()

    print(f"[engine] Saved base workbook: {os.path.basename(output_path)}")
    print(f"[engine] Tabs added: '{mtd_tab}', '{ytd_tab}'")

    # ── Open in Excel to recalculate all formulas ─────────────────
    _open_recalc_save(output_path)

    return output_path


# ═══════════════════════════════════════════════════════════════
#  FULL-FIDELITY SHEET COPY
# ═══════════════════════════════════════════════════════════════

def _copy_sheet_into(src_wb, src_ws, dst_wb, new_name: str):
    """
    Copies src_ws into dst_wb as a new sheet called new_name.

    Preserves:
      - Cell values and formulas
      - Font, fill, border, alignment, number format
      - Merged cell ranges
      - Column widths and row heights
      - Sheet view (grid lines, zoom, freeze panes)
      - Tab colour
      - Hyperlinks (internal and external — no breakage)
      - Data validations
      - Conditional formatting rules
      - Print / page-setup settings

    Note: Charts, images, sparklines cannot be reliably copied by openpyxl
          and are skipped silently.
    """
    dst_ws = dst_wb.create_sheet(new_name)

    # Sheet view
    dst_ws.sheet_view.showGridLines = src_ws.sheet_view.showGridLines
    dst_ws.sheet_view.zoomScale     = src_ws.sheet_view.zoomScale

    # Tab colour
    if src_ws.sheet_properties.tabColor:
        dst_ws.sheet_properties.tabColor = src_ws.sheet_properties.tabColor

    # Page setup
    for attr in ("orientation", "paperSize", "fitToPage",
                 "fitToHeight", "fitToWidth"):
        try:
            setattr(dst_ws.page_setup, attr,
                    getattr(src_ws.page_setup, attr))
        except Exception:
            pass

    # Column widths
    for col_letter, col_dim in src_ws.column_dimensions.items():
        dst_ws.column_dimensions[col_letter].width   = col_dim.width
        dst_ws.column_dimensions[col_letter].hidden  = col_dim.hidden
        dst_ws.column_dimensions[col_letter].bestFit = col_dim.bestFit

    # Row heights
    for row_idx, row_dim in src_ws.row_dimensions.items():
        dst_ws.row_dimensions[row_idx].height = row_dim.height
        dst_ws.row_dimensions[row_idx].hidden = row_dim.hidden

    # Cells — values, styles, hyperlinks
    for row in src_ws.iter_rows():
        for src_cell in row:
            dst_cell = dst_ws.cell(row=src_cell.row, column=src_cell.column)
            dst_cell.value         = src_cell.value
            dst_cell.number_format = src_cell.number_format
            if src_cell.font:
                dst_cell.font = copy.copy(src_cell.font)
            if src_cell.fill:
                dst_cell.fill = copy.copy(src_cell.fill)
            if src_cell.border:
                dst_cell.border = copy.copy(src_cell.border)
            if src_cell.alignment:
                dst_cell.alignment = copy.copy(src_cell.alignment)
            if src_cell.hyperlink:
                dst_cell.hyperlink = copy.copy(src_cell.hyperlink)

    # Merged cells
    for merge_range in src_ws.merged_cells.ranges:
        dst_ws.merge_cells(str(merge_range))

    # Data validations
    for dv in src_ws.data_validations.dataValidation:
        dst_ws.add_data_validation(copy.deepcopy(dv))

    # Conditional formatting
    for cf_range, cf_rules in src_ws.conditional_formatting._cf_rules.items():
        for rule in cf_rules:
            dst_ws.conditional_formatting.add(cf_range, copy.deepcopy(rule))

    # Freeze panes
    if src_ws.freeze_panes:
        dst_ws.freeze_panes = src_ws.freeze_panes

    return dst_ws


# ═══════════════════════════════════════════════════════════════
#  AUDIT LOG SHEET
# ═══════════════════════════════════════════════════════════════

def _write_input_log(ws, args, sc1, sc2):
    ws["A1"] = "INPUT FILE LOG"
    ws["A1"].font = Font(name="Helvetica", bold=True, size=12, color="C3002F")

    meta = [
        ("Generated",    datetime.now().strftime("%d %b %Y  %H:%M")),
        ("Scenario 1",   sc1),
        ("Scenario 2",   sc2),
        ("Pair",         f"{sc1} vs {sc2}"),
        ("Quarter",      args.get("quarter", "")),
        ("Month",        args.get("month", "")),
        ("Master File",  args.get("master_file", "")),
    ]
    lf = Font(name="Helvetica", size=9, bold=True, color="888888")
    vf = Font(name="Helvetica", size=9, color="C0C0C0")
    for i, (label, value) in enumerate(meta, start=2):
        ws.cell(row=i, column=1, value=label).font = lf
        ws.cell(row=i, column=2, value=value).font = vf

    # File list header
    row = len(meta) + 3
    ws.cell(row=row, column=1, value="INPUT FILES").font = Font(
        name="Helvetica", bold=True, size=9, color="C3002F")
    row += 1
    hf    = PatternFill("solid", fgColor="1C1C1C")
    hfont = Font(name="Helvetica", bold=True, color="C0C0C0", size=9)
    for col, h in enumerate(["#", "Folder", "File Name", "Full Path"], start=1):
        c = ws.cell(row=row, column=col, value=h)
        c.fill = hf
        c.font = hfont
    row += 1

    for i, fp in enumerate(args.get("input_files", []), start=1):
        ws.cell(row=row, column=1, value=i).font           = Font(size=8, color="666666")
        ws.cell(row=row, column=2, value=os.path.dirname(fp)).font  = Font(size=8, color="888888")
        ws.cell(row=row, column=3, value=os.path.basename(fp)).font = Font(size=8, color="C0C0C0")
        ws.cell(row=row, column=4, value=fp).font          = Font(size=8, color="444444")
        row += 1

    for col, width in zip("ABCD", [6, 40, 30, 60]):
        ws.column_dimensions[col].width = width
    ws.sheet_view.showGridLines = False


# ═══════════════════════════════════════════════════════════════
#  TEMPLATE SHEET AUTO-DETECT
# ═══════════════════════════════════════════════════════════════

def _resolve_template_sheets(tpl_path: str):
    """
    Returns (mtd_sheet_name, ytd_sheet_name).
    Uses module-level constants if set; otherwise auto-detects by
    scanning for 'MTD' and 'YTD' in sheet names (case-insensitive).
    """
    wb    = load_workbook(tpl_path, read_only=True, data_only=True)
    names = wb.sheetnames
    wb.close()

    mtd = MTD_SHEET_NAME
    ytd = YTD_SHEET_NAME

    if mtd is None:
        for n in names:
            if "mtd" in n.lower():
                mtd = n
                break
    if ytd is None:
        for n in names:
            if "ytd" in n.lower():
                ytd = n
                break

    if mtd is None:
        raise ValueError(
            f"No MTD sheet found in template.\nSheets: {names}\n"
            "Set MTD_SHEET_NAME at the top of variance_engine.py."
        )
    if ytd is None:
        raise ValueError(
            f"No YTD sheet found in template.\nSheets: {names}\n"
            "Set YTD_SHEET_NAME at the top of variance_engine.py."
        )
    return mtd, ytd


# ═══════════════════════════════════════════════════════════════
#  XLWINGS — open in Excel, recalculate, save & close
# ═══════════════════════════════════════════════════════════════

def _open_recalc_save(path: str):
    try:
        import xlwings as xw
        app = xw.App(visible=False, add_book=False)
        app.display_alerts  = False
        app.screen_updating = False
        wb = app.books.open(os.path.abspath(path))
        app.api.CalculateFull()
        wb.save()
        wb.close()
        app.quit()
        print(f"[engine] Excel recalc complete: {os.path.basename(path)}")
    except Exception as e:
        print(f"[engine] xlwings recalc skipped: {e}")


# ═══════════════════════════════════════════════════════════════
#  HELPERS
# ═══════════════════════════════════════════════════════════════

def _safe(s: str) -> str:
    return s.replace(" ", "_").replace("+", "p").replace("/", "-")


# ═══════════════════════════════════════════════════════════════
#  STANDALONE TEST
# ═══════════════════════════════════════════════════════════════

if __name__ == "__main__":
    test_args = {
        "scenario_1":    "FC 2+10",
        "scenario_2":    "Actuals",
        "quarter":       "Q1 (Apr-Jun)",
        "month":         "April",
        "input_folders": [r"C:\Test\InputData"],
        "input_files":   [
            r"C:\Test\InputData\Sales_Data.xlsx",
            r"C:\Test\InputData\Actuals_Upload.xlsx",
        ],
        "master_file":   r"C:\Test\InputData\Variance_Master.xlsx",
        "output_folder": r"C:\Test\Output",
        "timestamp":     datetime.now().strftime("%Y%m%d_%H%M%S"),
    }
    out = run_variance(test_args)
    print(f"Saved to: {out}")
