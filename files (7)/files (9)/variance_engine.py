"""
variance_engine.py
──────────────────
Called by the Nissan Variance App when the user clicks GENERATE FILE.

Logic:
  1.  Receive N scenario pairs  (e.g. [("FC 2+10","Actuals"), ("Budget","LY Actuals")])
  2.  Copy the user-selected master file to the output folder — this becomes
      the single output workbook.
  3.  Open the template file and auto-detect its MTD and YTD sheets.
  4.  For EVERY pair inject two tabs into the output workbook:
          MTD <sc1> vs <sc2>
          YTD <sc1> vs <sc2>
      All cell formatting, merges, freeze-panes, conditional formatting,
      column widths, row heights, hyperlinks and data-validations are preserved.
  5.  Append one "Input Log" audit sheet at the end.
  6.  Save.  Then optionally open in Excel to force a full formula recalc.

Receives a dict (args):
    args = {
        "scenario_pairs":  list[tuple[str,str]],   # [(sc1, sc2), ...]
        "quarter":         str,                     # "Q1 (Apr-Jun)"
        "month":           str,                     # "April"
        "input_folders":   list[str],
        "input_files":     list[str],
        "master_file":     str,                     # full path chosen in UI
        "template_file":   str,                     # full path chosen in UI
        "output_folder":   str,
        "timestamp":       str,                     # YYYYMMDD_HHMMSS
    }

Requirements:
    pip install openpyxl
    xlwings is optional — used only for Excel formula recalc (needs Excel installed).
"""

import os
import copy
import shutil
from datetime import datetime

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill


# ═══════════════════════════════════════════════════════════════
#  CONFIGURATION
# ═══════════════════════════════════════════════════════════════

# Set these to exact sheet names if auto-detect is unreliable.
# Leave as None to auto-detect by scanning for "MTD" / "YTD" in sheet names.
MTD_SHEET_NAME: str | None = None
YTD_SHEET_NAME: str | None = None


# ═══════════════════════════════════════════════════════════════
#  MAIN ENTRY POINT
# ═══════════════════════════════════════════════════════════════

def run_variance(args: dict) -> str:
    """
    Build one output workbook containing MTD + YTD tabs for every
    scenario pair in args["scenario_pairs"].

    Returns the full path of the saved output file.
    """
    pairs         = args["scenario_pairs"]          # [(sc1, sc2), ...]
    quarter       = args["quarter"]
    month         = args["month"]
    master_file   = args["master_file"]
    template_file = args["template_file"]
    out_folder    = args["output_folder"]
    ts            = args["timestamp"]

    os.makedirs(out_folder, exist_ok=True)

    # ── Validate source files ──────────────────────────────────
    if not os.path.isfile(template_file):
        raise FileNotFoundError(
            f"Template file not found:\n{template_file}\n\n"
            "Please browse for the correct template file in the UI."
        )
    if not os.path.isfile(master_file):
        raise FileNotFoundError(
            f"Master input file not found:\n{master_file}"
        )

    # ── Resolve MTD / YTD sheet names from template ────────────
    mtd_name, ytd_name = _resolve_template_sheets(template_file)

    # ── Build output file path ──────────────────────────────────
    # Name encodes the quarter + month only (pairs listed in audit log)
    q_label = quarter.split()[0]
    output_path = os.path.join(
        out_folder,
        f"Variance__{q_label}_{month}__v{ts}.xlsx"
    )

    # ── Copy master file → becomes the output workbook ─────────
    shutil.copy2(master_file, output_path)

    # ── Open both workbooks ─────────────────────────────────────
    out_wb  = load_workbook(output_path, data_only=False, keep_vba=False)
    tpl_wb  = load_workbook(template_file, data_only=False, keep_vba=False)

    mtd_src = tpl_wb[mtd_name]
    ytd_src = tpl_wb[ytd_name]

    # ── Inject MTD + YTD tabs for every pair ───────────────────
    inserted_tabs: list[str] = []

    for sc1, sc2 in pairs:
        pair_label = f"{sc1} vs {sc2}"
        mtd_tab    = f"MTD {pair_label}"[:31]
        ytd_tab    = f"YTD {pair_label}"[:31]

        # Remove any pre-existing tabs with the same name (idempotent re-run)
        for name in [mtd_tab, ytd_tab]:
            if name in out_wb.sheetnames:
                del out_wb[name]

        _copy_sheet_into(tpl_wb, mtd_src, out_wb, mtd_tab)
        _copy_sheet_into(tpl_wb, ytd_src, out_wb, ytd_tab)

        inserted_tabs += [mtd_tab, ytd_tab]
        print(f"[engine] Tabs added: '{mtd_tab}', '{ytd_tab}'")

    # ── Append audit log ────────────────────────────────────────
    if "Input Log" in out_wb.sheetnames:
        del out_wb["Input Log"]
    _write_input_log(out_wb.create_sheet("Input Log"), args, pairs)

    # ── Save ───────────────────────────────────────────────────
    out_wb.save(output_path)
    tpl_wb.close()
    print(f"[engine] Saved: {os.path.basename(output_path)}")

    # ── Optional Excel recalc ──────────────────────────────────
    _open_recalc_save(output_path)

    return output_path


# ═══════════════════════════════════════════════════════════════
#  FULL-FIDELITY SHEET COPY
# ═══════════════════════════════════════════════════════════════

def _copy_sheet_into(src_wb, src_ws, dst_wb, new_name: str):
    """
    Copy src_ws into dst_wb as a new sheet named new_name.

    Preserves:
      - Cell values and formulas
      - Font, fill, border, alignment, number format
      - Merged cell ranges
      - Column widths and row heights
      - Sheet view (grid lines, zoom)
      - Freeze panes
      - Tab colour
      - Hyperlinks
      - Data validations
      - Conditional formatting
      - Print / page-setup settings

    Note: Charts, images and sparklines cannot be reliably copied by
          openpyxl and are silently skipped.
    """
    dst_ws = dst_wb.create_sheet(new_name)

    # Sheet view
    dst_ws.sheet_view.showGridLines = src_ws.sheet_view.showGridLines
    dst_ws.sheet_view.zoomScale     = src_ws.sheet_view.zoomScale

    # Tab colour
    if src_ws.sheet_properties.tabColor:
        dst_ws.sheet_properties.tabColor = src_ws.sheet_properties.tabColor

    # Page setup
    for attr in ("orientation", "paperSize", "fitToPage",
                 "fitToHeight", "fitToWidth"):
        try:
            setattr(dst_ws.page_setup, attr,
                    getattr(src_ws.page_setup, attr))
        except Exception:
            pass

    # Column widths
    for col_letter, col_dim in src_ws.column_dimensions.items():
        dst_ws.column_dimensions[col_letter].width   = col_dim.width
        dst_ws.column_dimensions[col_letter].hidden  = col_dim.hidden
        dst_ws.column_dimensions[col_letter].bestFit = col_dim.bestFit

    # Row heights
    for row_idx, row_dim in src_ws.row_dimensions.items():
        dst_ws.row_dimensions[row_idx].height = row_dim.height
        dst_ws.row_dimensions[row_idx].hidden = row_dim.hidden

    # Cells — values, styles, hyperlinks
    for row in src_ws.iter_rows():
        for src_cell in row:
            dst_cell = dst_ws.cell(row=src_cell.row, column=src_cell.column)
            dst_cell.value         = src_cell.value
            dst_cell.number_format = src_cell.number_format
            if src_cell.font:
                dst_cell.font      = copy.copy(src_cell.font)
            if src_cell.fill:
                dst_cell.fill      = copy.copy(src_cell.fill)
            if src_cell.border:
                dst_cell.border    = copy.copy(src_cell.border)
            if src_cell.alignment:
                dst_cell.alignment = copy.copy(src_cell.alignment)
            if src_cell.hyperlink:
                dst_cell.hyperlink = copy.copy(src_cell.hyperlink)

    # Merged cells
    for merge_range in src_ws.merged_cells.ranges:
        dst_ws.merge_cells(str(merge_range))

    # Data validations
    for dv in src_ws.data_validations.dataValidation:
        dst_ws.add_data_validation(copy.deepcopy(dv))

    # Conditional formatting
    for cf_range, cf_rules in src_ws.conditional_formatting._cf_rules.items():
        for rule in cf_rules:
            dst_ws.conditional_formatting.add(cf_range, copy.deepcopy(rule))

    # Freeze panes
    if src_ws.freeze_panes:
        dst_ws.freeze_panes = src_ws.freeze_panes

    return dst_ws


# ═══════════════════════════════════════════════════════════════
#  AUDIT LOG SHEET
# ═══════════════════════════════════════════════════════════════

def _write_input_log(ws, args: dict, pairs: list):
    ws["A1"] = "INPUT FILE LOG"
    ws["A1"].font = Font(name="Helvetica", bold=True, size=12, color="C3002F")

    meta = [
        ("Generated",   datetime.now().strftime("%d %b %Y  %H:%M")),
        ("Quarter",     args.get("quarter", "")),
        ("Month",       args.get("month", "")),
        ("Master File", args.get("master_file", "")),
        ("Template",    args.get("template_file", "")),
    ]
    lf = Font(name="Helvetica", size=9, bold=True, color="888888")
    vf = Font(name="Helvetica", size=9, color="C0C0C0")
    for i, (label, value) in enumerate(meta, start=2):
        ws.cell(row=i, column=1, value=label).font = lf
        ws.cell(row=i, column=2, value=value).font = vf

    # Pairs table
    row = len(meta) + 3
    ws.cell(row=row, column=1, value="SCENARIO PAIRS").font = Font(
        name="Helvetica", bold=True, size=9, color="C3002F")
    row += 1
    hf    = PatternFill("solid", fgColor="1C1C1C")
    hfont = Font(name="Helvetica", bold=True, color="C0C0C0", size=9)
    for col, h in enumerate(["#", "Scenario 1", "Scenario 2", "MTD Tab", "YTD Tab"], start=1):
        c = ws.cell(row=row, column=col, value=h)
        c.fill = hf
        c.font = hfont
    row += 1

    for idx, (sc1, sc2) in enumerate(pairs, start=1):
        pair_label = f"{sc1} vs {sc2}"
        ws.cell(row=row, column=1, value=idx).font             = Font(size=8, color="666666")
        ws.cell(row=row, column=2, value=sc1).font             = Font(size=8, color="C0C0C0")
        ws.cell(row=row, column=3, value=sc2).font             = Font(size=8, color="C0C0C0")
        ws.cell(row=row, column=4, value=f"MTD {pair_label}"[:31]).font = Font(size=8, color="888888")
        ws.cell(row=row, column=5, value=f"YTD {pair_label}"[:31]).font = Font(size=8, color="888888")
        row += 1

    # Input files
    row += 1
    ws.cell(row=row, column=1, value="INPUT FILES").font = Font(
        name="Helvetica", bold=True, size=9, color="C3002F")
    row += 1
    for col, h in enumerate(["#", "Folder", "File Name", "Full Path"], start=1):
        c = ws.cell(row=row, column=col, value=h)
        c.fill = hf
        c.font = hfont
    row += 1

    for i, fp in enumerate(args.get("input_files", []), start=1):
        ws.cell(row=row, column=1, value=i).font             = Font(size=8, color="666666")
        ws.cell(row=row, column=2, value=os.path.dirname(fp)).font  = Font(size=8, color="888888")
        ws.cell(row=row, column=3, value=os.path.basename(fp)).font = Font(size=8, color="C0C0C0")
        ws.cell(row=row, column=4, value=fp).font            = Font(size=8, color="444444")
        row += 1

    for col, width in zip("ABCDE", [6, 20, 20, 30, 40]):
        ws.column_dimensions[col].width = width
    ws.sheet_view.showGridLines = False


# ═══════════════════════════════════════════════════════════════
#  TEMPLATE SHEET AUTO-DETECT
# ═══════════════════════════════════════════════════════════════

def _resolve_template_sheets(tpl_path: str):
    """
    Returns (mtd_sheet_name, ytd_sheet_name).
    Uses module-level constants if set; otherwise auto-detects.
    """
    wb    = load_workbook(tpl_path, read_only=True, data_only=True)
    names = wb.sheetnames
    wb.close()

    mtd = MTD_SHEET_NAME
    ytd = YTD_SHEET_NAME

    if mtd is None:
        for n in names:
            if "mtd" in n.lower():
                mtd = n
                break
    if ytd is None:
        for n in names:
            if "ytd" in n.lower():
                ytd = n
                break

    if mtd is None:
        raise ValueError(
            f"No MTD sheet found in template.\nSheets: {names}\n"
            "Set MTD_SHEET_NAME at the top of variance_engine.py."
        )
    if ytd is None:
        raise ValueError(
            f"No YTD sheet found in template.\nSheets: {names}\n"
            "Set YTD_SHEET_NAME at the top of variance_engine.py."
        )
    return mtd, ytd


# ═══════════════════════════════════════════════════════════════
#  XLWINGS — open in Excel, recalculate, save & close
# ═══════════════════════════════════════════════════════════════

def _open_recalc_save(path: str):
    """Open in Excel for a full recalc; silently skipped if xlwings unavailable."""
    try:
        import xlwings as xw
        app = xw.App(visible=False, add_book=False)
        app.display_alerts  = False
        app.screen_updating = False
        wb = app.books.open(os.path.abspath(path))
        app.api.CalculateFull()
        wb.save()
        wb.close()
        app.quit()
        print(f"[engine] Excel recalc complete: {os.path.basename(path)}")
    except Exception as e:
        print(f"[engine] xlwings recalc skipped: {e}")


# ═══════════════════════════════════════════════════════════════
#  HELPERS
# ═══════════════════════════════════════════════════════════════

def _safe(s: str) -> str:
    return s.replace(" ", "_").replace("+", "p").replace("/", "-")


# ═══════════════════════════════════════════════════════════════
#  STANDALONE TEST
# ═══════════════════════════════════════════════════════════════

if __name__ == "__main__":
    from datetime import datetime as _dt
    test_args = {
        "scenario_pairs":  [("FC 2+10", "Actuals"), ("Budget", "LY Actuals")],
        "quarter":         "Q1 (Apr-Jun)",
        "month":           "April",
        "input_folders":   [r"C:\Test\InputData"],
        "input_files":     [r"C:\Test\InputData\Sales_Data.xlsx"],
        "master_file":     r"C:\Test\InputData\Variance_Master.xlsx",
        "template_file":   r"C:\Test\Templates\Variance_Template.xlsx",
        "output_folder":   r"C:\Test\Output",
        "timestamp":       _dt.now().strftime("%Y%m%d_%H%M%S"),
    }
    out = run_variance(test_args)
    print(f"Saved to: {out}")
